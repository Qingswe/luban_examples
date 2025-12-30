import openpyxl
from openpyxl import load_workbook
import os

# 设置路径
data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
enums_file = os.path.join(data_dir, '__enums__.xlsx')
item_file = os.path.join(data_dir, '#demo.item.xlsx')

# Task 2.1: 物品基础表 (TbItem)

# 1. 在 __enums__.xlsx 中添加 Quality 和 ItemType 枚举
print("正在更新 __enums__.xlsx...")
wb_enums = load_workbook(enums_file)
sheet_enums = wb_enums.active

# 找到最后一行
last_row = sheet_enums.max_row

# 添加 Quality 枚举
enums_data = [
    ['', 'Quality', 'FALSE', 'TRUE', '', '物品品质', '', 'COMMON', '普通', '1', '普通品质'],
    ['', '', '', '', '', '', '', 'RARE', '稀有', '2', '稀有品质'],
    ['', '', '', '', '', '', '', 'EPIC', '史诗', '3', '史诗品质'],
    ['', '', '', '', '', '', '', 'LEGENDARY', '传说', '4', '传说品质'],
    ['', 'ItemType', 'FALSE', 'TRUE', '', '物品类型', '', 'CONSUMABLE', '消耗品', '1', '消耗品类型'],
    ['', '', '', '', '', '', '', 'MATERIAL', '材料', '2', '材料类型'],
    ['', '', '', '', '', '', '', 'EQUIPMENT', '装备', '3', '装备类型'],
]

for row_data in enums_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_enums.cell(last_row, col_idx).value = value

wb_enums.save(enums_file)
print(f"已更新 __enums__.xlsx，添加了 Quality 和 ItemType 枚举")

# 2. 创建 #demo.item.xlsx 数据表
print("正在创建 #demo.item.xlsx...")
wb_item = openpyxl.Workbook()
sheet = wb_item.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'desc'
sheet.cell(1, 5).value = 'quality'
sheet.cell(1, 6).value = 'item_type'
sheet.cell(1, 7).value = 'stack_limit'
sheet.cell(1, 8).value = 'sell_price'

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'text'
sheet.cell(2, 5).value = 'Quality'
sheet.cell(2, 6).value = 'ItemType'
sheet.cell(2, 7).value = 'int#range=[1,9999]'
sheet.cell(2, 8).value = 'int'

# 第3行: ##group (可选)
sheet.cell(3, 1).value = '##group'
sheet.cell(3, 2).value = ''
sheet.cell(3, 3).value = 'c,s'
sheet.cell(3, 4).value = 'c,s'
sheet.cell(3, 5).value = 'c,s'
sheet.cell(3, 6).value = 'c,s'
sheet.cell(3, 7).value = 'c,s'
sheet.cell(3, 8).value = 's'

# 第4行: ## (注释行)
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'
sheet.cell(4, 4).value = '描述'
sheet.cell(4, 5).value = '品质'
sheet.cell(4, 6).value = '类型'
sheet.cell(4, 7).value = '堆叠上限'
sheet.cell(4, 8).value = '售价'

# 数据行
items_data = [
    ['', 1001, 'item_health_potion', 'desc_health_potion', 'COMMON', 'CONSUMABLE', 99, 10],
    ['', 1002, 'item_mana_potion', 'desc_mana_potion', 'COMMON', 'CONSUMABLE', 99, 15],
    ['', 2001, 'item_iron_ore', 'desc_iron_ore', 'RARE', 'MATERIAL', 999, 5],
    ['', 2002, 'item_gold_ore', 'desc_gold_ore', 'EPIC', 'MATERIAL', 999, 50],
    ['', 3001, 'item_excalibur', 'desc_excalibur', 'LEGENDARY', 'EQUIPMENT', 1, 10000],
    ['', 3002, 'item_steel_sword', 'desc_steel_sword', 'RARE', 'EQUIPMENT', 1, 500],
]

for row_data in items_data:
    sheet.append(row_data)

wb_item.save(item_file)
print(f"已创建 #demo.item.xlsx")
print("\nTask 2.1 完成!")
print("创建的文件:")
print(f"  - 更新了 __enums__.xlsx (添加 Quality 和 ItemType)")
print(f"  - 创建了 #demo.item.xlsx (包含质量、类型、描述等字段)")
