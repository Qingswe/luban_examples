import openpyxl
from openpyxl import load_workbook

# 读取当前的__beans__.xlsx看看有什么问题
cur_beans = r'e:\Learn\luban_examples\DataTables\Datas\__beans__.xlsx'

wb = load_workbook(cur_beans)
sheet = wb.active

print("当前 __beans__.xlsx 内容 (前50行):")
print("=" * 120)

# 显示前50行
for row_idx in range(1, min(51, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, 8):  # 显示前7列
        cell_value = sheet.cell(row_idx, col_idx).value
        if cell_value is None:
            cell_value = ''
        row_data.append(str(cell_value)[:15])  # 限制每列最多15个字符

    # 标记问题行
    full_name = sheet.cell(row_idx, 2).value
    parent = sheet.cell(row_idx, 3).value
    field_name = sheet.cell(row_idx, 4).value

    marker = "  "
    if full_name == '' and parent == '' and field_name != '' and field_name != 'name' and row_idx > 3:
        marker = "❌"

    print(f"{marker} Row {row_idx:3d}: {' | '.join(row_data)}")

print("\n" + "=" * 120)
print(f"总行数: {sheet.max_row}")

# 找出问题行（第20行）
print(f"\n第20行的内容:")
for col_idx in range(1, 10):
    val = sheet.cell(20, col_idx).value
    print(f"  Col {col_idx}: '{val}'")
