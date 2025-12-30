import openpyxl
from openpyxl import load_workbook
import os
import shutil

# 重新生成正确的__beans__.xlsx

data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
beans_file = os.path.join(data_dir, '__beans__.xlsx')
backup_file = os.path.join(data_dir, '__beans_old_backup.xlsx')

# 备份原文件
shutil.copy(beans_file, backup_file)
print(f"已备份原文件到: {backup_file}")

# 加载原始文件以保留原有内容
wb = load_workbook(beans_file)
sheet = wb.active

# 获取标题行（前3行）
header_row1 = [sheet.cell(1, i).value for i in range(1, 8)]
header_row2 = [sheet.cell(2, i).value for i in range(1, 8)]
header_row3 = [sheet.cell(3, i).value for i in range(1, 8)]

# 创建新的workbook
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

# 写入标题行
for col_idx, val in enumerate(header_row1, start=1):
    sheet_new.cell(1, col_idx).value = val

# 第2行保持空值
for col_idx, val in enumerate(header_row2, start=1):
    sheet_new.cell(2, col_idx).value = val

# 第3行注释行
for col_idx, val in enumerate(header_row3, start=1):
    sheet_new.cell(3, col_idx).value = val

current_row = 4

# 定义所有的Bean（使用正确的格式）
# 注意：对于多态类型，字段应该在*fields列中定义

beans_definitions = [
    # PassiveEffect (基类，无字段)
    ('PassiveEffect', '', '', '被动效果基类'),

    # StatBonus (继承PassiveEffect)
    ('StatBonus', 'PassiveEffect', '', '属性加成'),
    ('', '', 'stat|string||属性名', ''),
    ('', '', 'value|int||数值', ''),
    ('', '', 'is_percent|bool||是否百分比', ''),

    # SkillEnhance (继承PassiveEffect)
    ('SkillEnhance', 'PassiveEffect', '', '技能强化'),
    ('', '', 'skill_id|int#ref=TbSkillTree||技能ID', ''),
    ('', '', 'bonus_damage|int||额外伤害', ''),

    # TalentNode
    ('TalentNode', '', '', '天赋节点'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|text||名称', ''),
    ('', '', 'tier|int||层级', ''),
    ('', '', '*effects|list,PassiveEffect||效果列表', ''),

    # Modifier (基类)
    ('Modifier', '', '', '属性调整器基类'),

    # FlatModifier
    ('FlatModifier', 'Modifier', '', '固定值调整器'),
    ('', '', 'stat|string||属性名', ''),
    ('', '', 'value|int||数值', ''),

    # PercentModifier
    ('PercentModifier', 'Modifier', '', '百分比调整器'),
    ('', '', 'stat|string||属性名', ''),
    ('', '', 'percent|float||百分比', ''),

    # Affix
    ('Affix', '', '', '词缀'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|string||名称', ''),
    ('', '', 'modifier|Modifier||调整器', ''),

    # Equipment
    ('Equipment', '', '', '装备'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|string||名称', ''),
    ('', '', 'slot|EquipSlot||装备槽位', ''),
    ('', '', 'base_stats|map,string,int||基础属性', ''),
    ('', '', 'affix_pool|(list#index=id),Affix||词缀池', ''),
    ('', '', 'set_id|int?||套装ID', ''),

    # Material
    ('Material', '', '', '材料'),
    ('', '', 'item_id|int#ref=TbItem||物品ID', ''),
    ('', '', 'count|int#range=[1,999]||数量', ''),

    # Recipe
    ('Recipe', '', '', '配方'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|string||名称', ''),
    ('', '', 'output_item|int#ref=TbItem||输出物品', ''),
    ('', '', 'output_count|int#range=[1,99]||输出数量', ''),
    ('', '', '*materials|list,Material||材料列表', ''),
    ('', '', 'unlock_level|int#range=[1,100]||解锁等级', ''),

    # Effect (基类)
    ('Effect', '', '', '效果基类'),

    # DamageEffect (继承Effect)
    ('DamageEffect', 'Effect', '', '伤害效果'),

    # PhysicalDamage (继承DamageEffect)
    ('PhysicalDamage', 'DamageEffect', '', '物理伤害'),
    ('', '', 'base_damage|int||基础伤害', ''),
    ('', '', 'armor_pen|float||护甲穿透', ''),

    # MagicalDamage (继承DamageEffect)
    ('MagicalDamage', 'DamageEffect', '', '魔法伤害'),
    ('', '', 'base_damage|int||基础伤害', ''),
    ('', '', 'element|Element||元素', ''),
    ('', '', 'ignore_resist|bool||无视抗性', ''),

    # HealEffect (继承Effect)
    ('HealEffect', 'Effect', '', '治疗效果'),
    ('', '', 'heal_base|int||基础治疗', ''),
    ('', '', 'heal_scale|float||治疗系数', ''),

    # SkillEffectConfig
    ('SkillEffectConfig', '', '', '技能效果配置'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|string||名称', ''),
    ('', '', 'effect|Effect||效果', ''),
    ('', '', 'scaling_stat|string||成长属性', ''),
    ('', '', 'scaling_factor|float||成长系数', ''),

    # BuffEffect (基类)
    ('BuffEffect', '', '', 'Buff效果基类'),

    # StatModifier (继承BuffEffect)
    ('StatModifier', 'BuffEffect', '', '属性调整'),
    ('', '', 'stat|string||属性名', ''),
    ('', '', 'value|int||数值', ''),
    ('', '', 'is_percent|bool||是否百分比', ''),

    # DotEffect (继承BuffEffect)
    ('DotEffect', 'BuffEffect', '', '持续伤害'),
    ('', '', 'damage_per_tick|int||每跳伤害', ''),
    ('', '', 'tick_interval|float||跳动间隔', ''),

    # ControlEffect (继承BuffEffect)
    ('ControlEffect', 'BuffEffect', '', '控制效果'),
    ('', '', 'control_type|string||控制类型', ''),

    # Buff
    ('Buff', '', '', 'Buff'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|text||名称', ''),
    ('', '', 'buff_type|BuffType||Buff类型', ''),
    ('', '', 'duration|float#range=[0,3600]||持续时间', ''),
    ('', '', 'stack_limit|int#range=[1,99]||叠加上限', ''),
    ('', '', 'refresh_on_apply|bool||重新施加时刷新', ''),
    ('', '', 'effect|BuffEffect||效果', ''),

    # ElementRelation
    ('ElementRelation', '', '', '元素克制关系'),
    ('', '', 'source_element|Element||源元素', ''),
    ('', '', 'damage_multipliers|map,Element,float||伤害倍率', ''),

    # Quest
    ('Quest', '', '', '任务'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'title|text||标题', ''),
    ('', '', 'desc|text||描述', ''),
    ('', '', 'quest_type|QuestType||任务类型', ''),
    ('', '', 'prereq_quest|int#ref=TbQuest?||前置任务', ''),
    ('', '', 'reward_gold|int||金币奖励', ''),
    ('', '', 'reward_exp|int||经验奖励', ''),

    # DialogChoice (基类)
    ('DialogChoice', '', '', '对话选项基类'),

    # SimpleChoice
    ('SimpleChoice', 'DialogChoice', '', '简单选项'),
    ('', '', 'text|text||文本', ''),
    ('', '', 'next_dialog|int#ref=TbDialog?||下一对话', ''),

    # ConditionChoice
    ('ConditionChoice', 'DialogChoice', '', '条件选项'),
    ('', '', 'text|text||文本', ''),
    ('', '', 'condition|string||条件', ''),
    ('', '', 'next_dialog|int#ref=TbDialog?||下一对话', ''),

    # ActionChoice
    ('ActionChoice', 'DialogChoice', '', '动作选项'),
    ('', '', 'text|text||文本', ''),
    ('', '', 'action|string||动作', ''),
    ('', '', 'params|string||参数', ''),

    # DialogLine
    ('DialogLine', '', '', '对话行'),
    ('', '', 'speaker|text||说话者', ''),
    ('', '', 'content|text||内容', ''),

    # Dialog
    ('Dialog', '', '', '对话'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|string||名称', ''),
    ('', '', '*lines|list,DialogLine||对话行', ''),
    ('', '', 'choices|list,DialogChoice||选项', ''),

    # AchievementCondition (基类)
    ('AchievementCondition', '', '', '成就条件基类'),

    # KillCount
    ('KillCount', 'AchievementCondition', '', '击杀数量'),
    ('', '', 'monster_type|string||怪物类型', ''),
    ('', '', 'count|int||数量', ''),

    # CollectItem
    ('CollectItem', 'AchievementCondition', '', '收集物品'),
    ('', '', 'item_id|int#ref=TbItem||物品ID', ''),
    ('', '', 'count|int||数量', ''),

    # CompleteQuest
    ('CompleteQuest', 'AchievementCondition', '', '完成任务'),
    ('', '', 'quest_id|int#ref=TbQuest||任务ID', ''),

    # ReachLevel
    ('ReachLevel', 'AchievementCondition', '', '达到等级'),
    ('', '', 'level|int#range=[1,100]||等级', ''),

    # AchievementReward
    ('AchievementReward', '', '', '成就奖励'),
    ('', '', 'item_id|int#ref=TbItem||物品ID', ''),
    ('', '', 'count|int#range=[1,999]||数量', ''),

    # Achievement
    ('Achievement', '', '', '成就'),
    ('', '', 'id|int||ID', ''),
    ('', '', 'name|text||名称', ''),
    ('', '', 'desc|text||描述', ''),
    ('', '', 'condition|AchievementCondition||条件', ''),
    ('', '', '*rewards|list,AchievementReward||奖励', ''),
    ('', '', 'points|int||成就点数', ''),
]

# 写入Bean定义
for bean_def in beans_definitions:
    full_name, parent, fields, comment = bean_def

    if full_name:  # Bean定义行
        sheet_new.cell(current_row, 1).value = ''  # ##var列
        sheet_new.cell(current_row, 2).value = full_name
        sheet_new.cell(current_row, 3).value = parent
        sheet_new.cell(current_row, 4).value = ''  # *fields列
        sheet_new.cell(current_row, 5).value = ''  # sep列
        sheet_new.cell(current_row, 6).value = ''  # alias列
        sheet_new.cell(current_row, 7).value = comment
    else:  # 字段定义行
        sheet_new.cell(current_row, 1).value = ''  # ##var列
        sheet_new.cell(current_row, 2).value = ''  # full_name列
        sheet_new.cell(current_row, 3).value = ''  # parent列

        # 解析fields字符串: name|type|group|comment
        if '|' in fields:
            parts = fields.split('|')
            field_name = parts[0] if len(parts) > 0 else ''
            field_type = parts[1] if len(parts) > 1 else ''
            field_group = parts[2] if len(parts) > 2 else ''
            field_comment = parts[3] if len(parts) > 3 else ''

            # *fields是多行列表，所以第4列应该是: name, type, group, comment 的子行
            # 但在子行中，我们需要通过##var子行来标识
            # 实际上，在__beans__.xlsx中，*fields应该在标题头中标记
            # 让我参考文档中的格式

            # 根据SKILL.md的示例，字段行的格式应该是:
            # ['', '', '', 'name', 'type', 'group', 'comment']
            # 即：前三列为空，然后是4个字段属性列
            sheet_new.cell(current_row, 4).value = field_name    # *fields中的name
            sheet_new.cell(current_row, 5).value = field_type    # *fields中的type
            sheet_new.cell(current_row, 6).value = field_group   # *fields中的group
            sheet_new.cell(current_row, 7).value = field_comment # *fields中的comment

    current_row += 1

# 保存新文件
wb_new.save(beans_file)
print(f"已重新生成 __beans__.xlsx")
print(f"总共写入 {current_row - 4} 行Bean定义")
print("\n完成！")
