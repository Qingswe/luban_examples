import openpyxl
from openpyxl import load_workbook

# 读取原始模板来获取正确的列结构
mini_beans = r'e:\Learn\luban_examples\MiniTemplate\Datas\__beans__.xlsx'
wb_mini = load_workbook(mini_beans)
sheet_mini = wb_mini.active

print("MiniTemplate __beans__.xlsx 标题行:")
for col_idx in range(1, 15):
    val = sheet_mini.cell(1, col_idx).value
    if val:
        print(f"  Col {col_idx}: {val}")

# 读取当前文件
cur_beans = r'e:\Learn\luban_examples\DataTables\Datas\__beans__.xlsx'
wb_cur = load_workbook(cur_beans)
sheet_cur = wb_cur.active

print("\n当前 __beans__.xlsx 标题行:")
for col_idx in range(1, 15):
    val = sheet_cur.cell(1, col_idx).value
    if val:
        print(f"  Col {col_idx}: {val}")
