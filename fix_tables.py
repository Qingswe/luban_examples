import openpyxl
from openpyxl import load_workbook
import os

# 修复__tables__.xlsx中的read_schema_from_file设置

data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
tables_file = os.path.join(data_dir, '__tables__.xlsx')

wb = load_workbook(tables_file)
sheet = wb.active

print("检查__tables__.xlsx中的注册信息...")

# 遍历所有行，找到equipment.xlsx、buff.xlsx、quest.xlsx的注册
for row_idx in range(2, sheet.max_row + 1):
    table_name = sheet.cell(row_idx, 2).value
    value_type = sheet.cell(row_idx, 3).value
    read_from_file = sheet.cell(row_idx, 4).value
    input_file = sheet.cell(row_idx, 5).value

    if table_name in ['TbEquipment', 'TbBuff', 'TbQuest']:
        print(f"\nRow {row_idx}: {table_name}")
        print(f"  value_type: {value_type}")
        print(f"  read_schema_from_file: {read_from_file}")
        print(f"  input: {input_file}")

        # 这些表的schema已经在__beans__.xlsx中定义了
        # 所以read_schema_from_file应该设置为FALSE
        if read_from_file == 'TRUE' or read_from_file == True:
            print(f"  -> 修改为 FALSE (因为{value_type}已在__beans__.xlsx中定义)")
            sheet.cell(row_idx, 4).value = 'FALSE'

wb.save(tables_file)
print("\n已修复__tables__.xlsx")
