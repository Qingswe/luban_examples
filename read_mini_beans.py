import openpyxl
from openpyxl import load_workbook

# 读取原始模板的__beans__.xlsx来了解格式
mini_beans = r'e:\Learn\luban_examples\MiniTemplate\Datas\__beans__.xlsx'

wb = load_workbook(mini_beans)
sheet = wb.active

print("MiniTemplate __beans__.xlsx 内容:")
print("=" * 80)

# 显示前20行
for row_idx in range(1, min(21, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, 8):  # 显示前7列
        cell_value = sheet.cell(row_idx, col_idx).value
        if cell_value is None:
            cell_value = ''
        row_data.append(str(cell_value))

    print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

print("\n" + "=" * 80)
print(f"总行数: {sheet.max_row}")
