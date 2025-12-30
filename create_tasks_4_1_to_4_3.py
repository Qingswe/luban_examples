import openpyxl
from openpyxl import load_workbook
import os

# 设置路径
data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
enums_file = os.path.join(data_dir, '__enums__.xlsx')
beans_file = os.path.join(data_dir, '__beans__.xlsx')
tables_file = os.path.join(data_dir, '__tables__.xlsx')

print("=" * 60)
print("开始执行最后的测试任务 (Task 4.1 ~ Task 4.3)")
print("=" * 60)

# ============================================================
# Task 4.1: 任务表 (TbQuest)
# ============================================================
print("\n[Task 4.1] 任务表 (TbQuest)")
print("-" * 60)

# 1. 添加 QuestType 枚举
wb_enums = load_workbook(enums_file)
sheet_enums = wb_enums.active
last_row = sheet_enums.max_row

enums_data = [
    ['', 'QuestType', 'FALSE', 'TRUE', '', '任务类型', '', 'MAIN', '主线', '1', '主线任务'],
    ['', '', '', '', '', '', '', 'SIDE', '支线', '2', '支线任务'],
    ['', '', '', '', '', '', '', 'DAILY', '日常', '3', '日常任务'],
    ['', '', '', '', '', '', '', 'EVENT', '活动', '4', '活动任务'],
]

for row_data in enums_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_enums.cell(last_row, col_idx).value = value

wb_enums.save(enums_file)
print("[OK] 已添加 QuestType 枚举")

# 2. 添加 Quest Bean (带变体)
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'Quest', '', '', '', '', '任务'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'title', 'text', '', '标题'],
    ['', '', '', 'desc', 'text', '', '描述'],
    ['', '', '', 'quest_type', 'QuestType', '', '任务类型'],
    ['', '', '', 'prereq_quest', 'int#ref=TbQuest?', '', '前置任务'],
    ['', '', '', 'reward_gold', 'int', '', '金币奖励'],
    ['', '', '', 'reward_exp', 'int', '', '经验奖励'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 Quest Bean")

# 3. 在 __tables__.xlsx 中注册 TbQuest
wb_tables = load_workbook(tables_file)
sheet_tables = wb_tables.active
sheet_tables.append(['', 'TbQuest', 'Quest', 'TRUE', 'quest.xlsx', '', ''])
wb_tables.save(tables_file)
print("[OK] 已在 __tables__.xlsx 中注册 TbQuest")

# 4. 创建 quest.xlsx
quest_file = os.path.join(data_dir, 'quest.xlsx')
wb_quest = openpyxl.Workbook()
sheet = wb_quest.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'title'
sheet.cell(1, 4).value = 'desc'
sheet.cell(1, 5).value = 'quest_type'
sheet.cell(1, 6).value = 'prereq_quest'
sheet.cell(1, 7).value = 'reward_gold'
sheet.cell(1, 8).value = 'reward_exp'

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'text'
sheet.cell(2, 5).value = 'QuestType'
sheet.cell(2, 6).value = 'int#ref=TbQuest?'
sheet.cell(2, 7).value = 'int'
sheet.cell(2, 8).value = 'int'

sheet.cell(3, 1).value = '##'
sheet.cell(3, 2).value = 'ID'
sheet.cell(3, 3).value = '标题'
sheet.cell(3, 4).value = '描述'

# 数据
sheet.append(['', 8001, 'quest_intro', 'quest_intro_desc', 'MAIN', '', 100, 500])
sheet.append(['', 8002, 'quest_first_battle', 'quest_first_battle_desc', 'MAIN', 8001, 200, 1000])
sheet.append(['', 8101, 'quest_collect_herbs', 'quest_collect_herbs_desc', 'SIDE', '', 50, 300])
sheet.append(['', 8201, 'quest_daily_patrol', 'quest_daily_patrol_desc', 'DAILY', '', 30, 100])

wb_quest.save(quest_file)
print("[OK] 已创建 quest.xlsx")

# ============================================================
# Task 4.2: 对话表 (TbDialog)
# ============================================================
print("\n[Task 4.2] 对话表 (TbDialog)")
print("-" * 60)

# 1. 添加 DialogChoice 多态、DialogLine、Dialog Bean
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'DialogChoice', '', '', '', '', '对话选项基类'],
    ['', 'SimpleChoice', 'DialogChoice', '', '', '', '简单选项'],
    ['', '', '', 'text', 'text', '', '文本'],
    ['', '', '', 'next_dialog', 'int#ref=TbDialog?', '', '下一对话'],
    ['', 'ConditionChoice', 'DialogChoice', '', '', '', '条件选项'],
    ['', '', '', 'text', 'text', '', '文本'],
    ['', '', '', 'condition', 'string', '', '条件'],
    ['', '', '', 'next_dialog', 'int#ref=TbDialog?', '', '下一对话'],
    ['', 'ActionChoice', 'DialogChoice', '', '', '', '动作选项'],
    ['', '', '', 'text', 'text', '', '文本'],
    ['', '', '', 'action', 'string', '', '动作'],
    ['', '', '', 'params', 'string', '', '参数'],
    ['', 'DialogLine', '', '', '', '', '对话行'],
    ['', '', '', 'speaker', 'text', '', '说话者'],
    ['', '', '', 'content', 'text', '', '内容'],
    ['', 'Dialog', '', '', '', '', '对话'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'string', '', '名称'],
    ['', '', '', '*lines', 'list,DialogLine', '', '对话行'],
    ['', '', '', 'choices', 'list,DialogChoice', '', '选项'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 DialogChoice 多态、DialogLine、Dialog Bean")

# 2. 创建 #Dialog.xlsx
dialog_file = os.path.join(data_dir, '#Dialog.xlsx')
wb_dialog = openpyxl.Workbook()
sheet = wb_dialog.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = '*lines'
sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
sheet.cell(1, 7).value = 'choices'
sheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=12)

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'list,DialogLine'
sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=6)
sheet.cell(2, 7).value = 'list,DialogChoice'
sheet.merge_cells(start_row=2, start_column=7, end_row=2, end_column=12)

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 4).value = 'speaker'
sheet.cell(3, 5).value = 'content'
sheet.cell(3, 7).value = '$type'
sheet.cell(3, 8).value = 'text'
sheet.cell(3, 9).value = 'next_dialog'
sheet.cell(3, 10).value = 'condition'
sheet.cell(3, 11).value = 'action'
sheet.cell(3, 12).value = 'params'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据: 对话1
row = 5
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 9001
sheet.cell(row, 3).value = 'intro_dialog'
sheet.cell(row, 4).value = 'npc_elder'
sheet.cell(row, 5).value = 'dialog_elder_greeting'
sheet.cell(row, 7).value = 'SimpleChoice'
sheet.cell(row, 8).value = 'choice_accept'
sheet.cell(row, 9).value = 9002

# 第二行对话
row = 6
sheet.cell(row, 4).value = 'npc_elder'
sheet.cell(row, 5).value = 'dialog_elder_explain'

# 第二个选项
row = 7
sheet.cell(row, 7).value = 'SimpleChoice'
sheet.cell(row, 8).value = 'choice_decline'
sheet.cell(row, 9).value = ''

wb_dialog.save(dialog_file)
print("[OK] 已创建 #Dialog.xlsx")

# ============================================================
# Task 4.3: 成就表 (TbAchievement)
# ============================================================
print("\n[Task 4.3] 成就表 (TbAchievement)")
print("-" * 60)

# 1. 添加 AchievementCondition 多态、AchievementReward、Achievement Bean
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'AchievementCondition', '', '', '', '', '成就条件基类'],
    ['', 'KillCount', 'AchievementCondition', '', '', '', '击杀数量'],
    ['', '', '', 'monster_type', 'string', '', '怪物类型'],
    ['', '', '', 'count', 'int', '', '数量'],
    ['', 'CollectItem', 'AchievementCondition', '', '', '', '收集物品'],
    ['', '', '', 'item_id', 'int#ref=TbItem', '', '物品ID'],
    ['', '', '', 'count', 'int', '', '数量'],
    ['', 'CompleteQuest', 'AchievementCondition', '', '', '', '完成任务'],
    ['', '', '', 'quest_id', 'int#ref=TbQuest', '', '任务ID'],
    ['', 'ReachLevel', 'AchievementCondition', '', '', '', '达到等级'],
    ['', '', '', 'level', 'int#range=[1,100]', '', '等级'],
    ['', 'AchievementReward', '', '', '', '', '成就奖励'],
    ['', '', '', 'item_id', 'int#ref=TbItem', '', '物品ID'],
    ['', '', '', 'count', 'int#range=[1,999]', '', '数量'],
    ['', 'Achievement', '', '', '', '', '成就'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'text', '', '名称'],
    ['', '', '', 'desc', 'text', '', '描述'],
    ['', '', '', 'condition', 'AchievementCondition', '', '条件'],
    ['', '', '', '*rewards', 'list,AchievementReward', '', '奖励'],
    ['', '', '', 'points', 'int', '', '成就点数'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 AchievementCondition 多态、AchievementReward、Achievement Bean")

# 2. 创建 #Achievement.xlsx
achievement_file = os.path.join(data_dir, '#Achievement.xlsx')
wb_ach = openpyxl.Workbook()
sheet = wb_ach.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'desc'
sheet.cell(1, 5).value = 'condition'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=9)
sheet.cell(1, 10).value = '*rewards'
sheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)
sheet.cell(1, 13).value = 'points'

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'text'
sheet.cell(2, 5).value = 'AchievementCondition'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=9)
sheet.cell(2, 10).value = 'list,AchievementReward'
sheet.merge_cells(start_row=2, start_column=10, end_row=2, end_column=12)
sheet.cell(2, 13).value = 'int'

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 5).value = '$type'
sheet.cell(3, 6).value = 'monster_type'
sheet.cell(3, 7).value = 'count'
sheet.cell(3, 8).value = 'item_id'
sheet.cell(3, 9).value = 'quest_id'
sheet.cell(3, 10).value = 'level'
sheet.cell(3, 11).value = 'item_id'
sheet.cell(3, 12).value = 'count'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据: 成就1 - 击杀怪物
row = 5
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 10001
sheet.cell(row, 3).value = 'ach_kill_goblin'
sheet.cell(row, 4).value = 'ach_kill_goblin_desc'
sheet.cell(row, 5).value = 'KillCount'
sheet.cell(row, 6).value = 'goblin'
sheet.cell(row, 7).value = 100
sheet.cell(row, 11).value = 1001
sheet.cell(row, 12).value = 10
sheet.cell(row, 13).value = 50

# 第二个奖励
row = 6
sheet.cell(row, 11).value = 1002
sheet.cell(row, 12).value = 5

# 成就2 - 收集物品
row = 7
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 10002
sheet.cell(row, 3).value = 'ach_collect_gold'
sheet.cell(row, 4).value = 'ach_collect_gold_desc'
sheet.cell(row, 5).value = 'CollectItem'
sheet.cell(row, 8).value = 2002
sheet.cell(row, 7).value = 50
sheet.cell(row, 11).value = 3001
sheet.cell(row, 12).value = 1
sheet.cell(row, 13).value = 100

# 成就3 - 完成任务
row = 8
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 10003
sheet.cell(row, 3).value = 'ach_main_story'
sheet.cell(row, 4).value = 'ach_main_story_desc'
sheet.cell(row, 5).value = 'CompleteQuest'
sheet.cell(row, 9).value = 8002
sheet.cell(row, 11).value = 1001
sheet.cell(row, 12).value = 50
sheet.cell(row, 13).value = 200

wb_ach.save(achievement_file)
print("[OK] 已创建 #Achievement.xlsx")

print("\n" + "=" * 60)
print("Task 4.1 ~ Task 4.3 全部完成!")
print("=" * 60)
