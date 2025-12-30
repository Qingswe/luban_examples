import openpyxl
from openpyxl import load_workbook
import os

# 设置路径
data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
enums_file = os.path.join(data_dir, '__enums__.xlsx')
beans_file = os.path.join(data_dir, '__beans__.xlsx')
tables_file = os.path.join(data_dir, '__tables__.xlsx')

print("=" * 60)
print("开始执行剩余测试任务 (Task 2.2 ~ Task 4.3)")
print("=" * 60)

# ============================================================
# Task 2.2: 装备表 (TbEquipment)
# ============================================================
print("\n[Task 2.2] 装备表 (TbEquipment)")
print("-" * 60)

# 1. 添加 EquipSlot 枚举
wb_enums = load_workbook(enums_file)
sheet_enums = wb_enums.active
last_row = sheet_enums.max_row

enums_data = [
    ['', 'EquipSlot', 'FALSE', 'TRUE', '', '装备槽位', '', 'WEAPON', '武器', '1', '武器槽'],
    ['', '', '', '', '', '', '', 'ARMOR', '护甲', '2', '护甲槽'],
    ['', '', '', '', '', '', '', 'HELMET', '头盔', '3', '头盔槽'],
    ['', '', '', '', '', '', '', 'BOOTS', '靴子', '4', '靴子槽'],
    ['', '', '', '', '', '', '', 'ACCESSORY', '饰品', '5', '饰品槽'],
]

for row_data in enums_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_enums.cell(last_row, col_idx).value = value

wb_enums.save(enums_file)
print("[OK] 已添加 EquipSlot 枚举")

# 2. 添加 Modifier 多态、Affix、Equipment Bean
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'Modifier', '', '', '', '', '属性调整器基类'],
    ['', 'FlatModifier', 'Modifier', '', '', '', '固定值调整器'],
    ['', '', '', 'stat', 'string', '', '属性名'],
    ['', '', '', 'value', 'int', '', '数值'],
    ['', 'PercentModifier', 'Modifier', '', '', '', '百分比调整器'],
    ['', '', '', 'stat', 'string', '', '属性名'],
    ['', '', '', 'percent', 'float', '', '百分比'],
    ['', 'Affix', '', '', '', '', '词缀'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'string', '', '名称'],
    ['', '', '', 'modifier', 'Modifier', '', '调整器'],
    ['', 'Equipment', '', '', '', '', '装备'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'string', '', '名称'],
    ['', '', '', 'slot', 'EquipSlot', '', '装备槽位'],
    ['', '', '', 'base_stats', 'map,string,int', '', '基础属性'],
    ['', '', '', 'affix_pool', '(list#index=id),Affix', '', '词缀池'],
    ['', '', '', 'set_id', 'int?', '', '套装ID'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 Modifier 多态、Affix、Equipment Bean")

# 3. 在 __tables__.xlsx 中注册 TbEquipment
wb_tables = load_workbook(tables_file)
sheet_tables = wb_tables.active
last_row = sheet_tables.max_row

sheet_tables.append(['', 'TbEquipment', 'Equipment', 'TRUE', 'equipment.xlsx', '', ''])
wb_tables.save(tables_file)
print("[OK] 已在 __tables__.xlsx 中注册 TbEquipment")

# 4. 创建 equipment.xlsx
equipment_file = os.path.join(data_dir, 'equipment.xlsx')
wb_eq = openpyxl.Workbook()
sheet = wb_eq.active

# 标题头
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'slot'
sheet.cell(1, 5).value = 'base_stats'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
sheet.cell(1, 8).value = 'affix_pool'
sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
sheet.cell(1, 14).value = 'set_id'

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'EquipSlot'
sheet.cell(2, 5).value = 'map,string,int'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
sheet.cell(2, 8).value = '(list#index=id),Affix'
sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)
sheet.cell(2, 14).value = 'int?'

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 5).value = 'hp'
sheet.cell(3, 6).value = 'atk'
sheet.cell(3, 7).value = 'def'
sheet.cell(3, 8).value = 'id'
sheet.cell(3, 9).value = 'name'
sheet.cell(3, 10).value = '$type'
sheet.cell(3, 11).value = 'stat'
sheet.cell(3, 12).value = 'value'
sheet.cell(3, 13).value = 'percent'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'
sheet.cell(4, 4).value = '槽位'
sheet.cell(4, 5).value = 'HP'
sheet.cell(4, 6).value = 'ATK'
sheet.cell(4, 7).value = 'DEF'

# 数据: 铁剑
sheet.append(['', 4001, 'Iron Sword', 'WEAPON', 0, 30, 5, 101, 'Sharp', 'FlatModifier', 'atk', 10, '', ''])

# 数据: 钢铠甲
sheet.append(['', 4002, 'Steel Armor', 'ARMOR', 100, 0, 50, 102, 'Tough', 'FlatModifier', 'def', 20, '', 1001])

wb_eq.save(equipment_file)
print("[OK] 已创建 equipment.xlsx")

# ============================================================
# Task 2.3: 合成配方表 (TbRecipe)
# ============================================================
print("\n[Task 2.3] 合成配方表 (TbRecipe)")
print("-" * 60)

# 1. 添加 Material 和 Recipe Bean
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'Material', '', '', '', '', '材料'],
    ['', '', '', 'item_id', 'int#ref=TbItem', '', '物品ID'],
    ['', '', '', 'count', 'int#range=[1,999]', '', '数量'],
    ['', 'Recipe', '', '', '', '', '配方'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'string', '', '名称'],
    ['', '', '', 'output_item', 'int#ref=TbItem', '', '输出物品'],
    ['', '', '', 'output_count', 'int#range=[1,99]', '', '输出数量'],
    ['', '', '', '*materials', 'list,Material', '', '材料列表'],
    ['', '', '', 'unlock_level', 'int#range=[1,100]', '', '解锁等级'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 Material 和 Recipe Bean")

# 2. 创建 #Recipe.xlsx
recipe_file = os.path.join(data_dir, '#Recipe.xlsx')
wb_recipe = openpyxl.Workbook()
sheet = wb_recipe.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'output_item'
sheet.cell(1, 5).value = 'output_count'
sheet.cell(1, 6).value = '*materials'
sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
sheet.cell(1, 9).value = 'unlock_level'

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'int#ref=TbItem'
sheet.cell(2, 5).value = 'int#range=[1,99]'
sheet.cell(2, 6).value = 'list,Material'
sheet.merge_cells(start_row=2, start_column=6, end_row=2, end_column=8)
sheet.cell(2, 9).value = 'int#range=[1,100]'

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 6).value = 'item_id'
sheet.cell(3, 7).value = 'count'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '配方名'

# 数据: 铁剑配方
row = 5
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 5001
sheet.cell(row, 3).value = 'Recipe Iron Sword'
sheet.cell(row, 4).value = 3002
sheet.cell(row, 5).value = 1
sheet.cell(row, 6).value = 2001
sheet.cell(row, 7).value = 10
sheet.cell(row, 9).value = 5

row = 6
sheet.cell(row, 6).value = 2002
sheet.cell(row, 7).value = 2

wb_recipe.save(recipe_file)
print("[OK] 已创建 #Recipe.xlsx")

# ============================================================
# Task 3.1: 技能效果表 (TbSkillEffect)
# ============================================================
print("\n[Task 3.1] 技能效果表 (TbSkillEffect)")
print("-" * 60)

# 1. 添加 Element 枚举
wb_enums = load_workbook(enums_file)
sheet_enums = wb_enums.active
last_row = sheet_enums.max_row

enums_data = [
    ['', 'Element', 'FALSE', 'TRUE', '', '元素类型', '', 'PHYSICAL', '物理', '0', ''],
    ['', '', '', '', '', '', '', 'FIRE', '火', '1', ''],
    ['', '', '', '', '', '', '', 'ICE', '冰', '2', ''],
    ['', '', '', '', '', '', '', 'LIGHTNING', '雷', '3', ''],
    ['', '', '', '', '', '', '', 'HOLY', '神圣', '4', ''],
    ['', '', '', '', '', '', '', 'DARK', '黑暗', '5', ''],
]

for row_data in enums_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_enums.cell(last_row, col_idx).value = value

wb_enums.save(enums_file)
print("[OK] 已添加 Element 枚举")

# 2. 添加 Effect 多态层次
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'Effect', '', '', '', '', '效果基类'],
    ['', 'DamageEffect', 'Effect', '', '', '', '伤害效果'],
    ['', 'PhysicalDamage', 'DamageEffect', '', '', '', '物理伤害'],
    ['', '', '', 'base_damage', 'int', '', '基础伤害'],
    ['', '', '', 'armor_pen', 'float', '', '护甲穿透'],
    ['', 'MagicalDamage', 'DamageEffect', '', '', '', '魔法伤害'],
    ['', '', '', 'base_damage', 'int', '', '基础伤害'],
    ['', '', '', 'element', 'Element', '', '元素'],
    ['', '', '', 'ignore_resist', 'bool', '', '无视抗性'],
    ['', 'HealEffect', 'Effect', '', '', '', '治疗效果'],
    ['', '', '', 'heal_base', 'int', '', '基础治疗'],
    ['', '', '', 'heal_scale', 'float', '', '治疗系数'],
    ['', 'SkillEffectConfig', '', '', '', '', '技能效果配置'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'string', '', '名称'],
    ['', '', '', 'effect', 'Effect', '', '效果'],
    ['', '', '', 'scaling_stat', 'string', '', '成长属性'],
    ['', '', '', 'scaling_factor', 'float', '', '成长系数'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 Effect 多态层次")

# 3. 创建 #SkillEffect.xlsx
skill_eff_file = os.path.join(data_dir, '#SkillEffect.xlsx')
wb_se = openpyxl.Workbook()
sheet = wb_se.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'effect'
sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=9)
sheet.cell(1, 10).value = 'scaling_stat'
sheet.cell(1, 11).value = 'scaling_factor'

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'Effect'
sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=9)
sheet.cell(2, 10).value = 'string'
sheet.cell(2, 11).value = 'float'

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 4).value = '$type'
sheet.cell(3, 5).value = 'base_damage'
sheet.cell(3, 6).value = 'armor_pen'
sheet.cell(3, 7).value = 'element'
sheet.cell(3, 8).value = 'ignore_resist'
sheet.cell(3, 9).value = 'heal_base'
sheet.cell(3, 10).value = 'heal_scale'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据
sheet.append(['', 6001, 'Physical Strike', 'PhysicalDamage', 100, 0.3, '', '', '', '', 'atk', 1.5])
sheet.append(['', 6002, 'Fire Blast', 'MagicalDamage', 150, '', 'FIRE', True, '', '', 'mp', 2.0])
sheet.append(['', 6003, 'Heal', 'HealEffect', '', '', '', '', 80, 1.2, 'mp', 1.0])

wb_se.save(skill_eff_file)
print("[OK] 已创建 #SkillEffect.xlsx")

# ============================================================
# Task 3.2: Buff表 (TbBuff)
# ============================================================
print("\n[Task 3.2] Buff表 (TbBuff)")
print("-" * 60)

# 1. 添加 BuffType 枚举 (flags)
wb_enums = load_workbook(enums_file)
sheet_enums = wb_enums.active
last_row = sheet_enums.max_row

enums_data = [
    ['', 'BuffType', 'TRUE', 'TRUE', '', 'Buff类型', '', 'NONE', '无', '0', ''],
    ['', '', '', '', '', '', '', 'POSITIVE', '正面', '1', ''],
    ['', '', '', '', '', '', '', 'NEGATIVE', '负面', '2', ''],
    ['', '', '', '', '', '', '', 'DISPELLABLE', '可驱散', '4', ''],
    ['', '', '', '', '', '', '', 'STACKABLE', '可叠加', '8', ''],
]

for row_data in enums_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_enums.cell(last_row, col_idx).value = value

wb_enums.save(enums_file)
print("[OK] 已添加 BuffType 枚举 (flags)")

# 2. 添加 BuffEffect 多态
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'BuffEffect', '', '', '', '', 'Buff效果基类'],
    ['', 'StatModifier', 'BuffEffect', '', '', '', '属性调整'],
    ['', '', '', 'stat', 'string', '', '属性名'],
    ['', '', '', 'value', 'int', '', '数值'],
    ['', '', '', 'is_percent', 'bool', '', '是否百分比'],
    ['', 'DotEffect', 'BuffEffect', '', '', '', '持续伤害'],
    ['', '', '', 'damage_per_tick', 'int', '', '每跳伤害'],
    ['', '', '', 'tick_interval', 'float', '', '跳动间隔'],
    ['', 'ControlEffect', 'BuffEffect', '', '', '', '控制效果'],
    ['', '', '', 'control_type', 'string', '', '控制类型'],
    ['', 'Buff', '', '', '', '', 'Buff'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'text', '', '名称'],
    ['', '', '', 'buff_type', 'BuffType', '', 'Buff类型'],
    ['', '', '', 'duration', 'float#range=[0,3600]', '', '持续时间'],
    ['', '', '', 'stack_limit', 'int#range=[1,99]', '', '叠加上限'],
    ['', '', '', 'refresh_on_apply', 'bool', '', '重新施加时刷新'],
    ['', '', '', 'effect', 'BuffEffect', '', '效果'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 BuffEffect 多态和 Buff Bean")

# 3. 在 __tables__.xlsx 中注册 TbBuff
wb_tables = load_workbook(tables_file)
sheet_tables = wb_tables.active
sheet_tables.append(['', 'TbBuff', 'Buff', 'TRUE', 'buff.xlsx', '', ''])
wb_tables.save(tables_file)
print("[OK] 已在 __tables__.xlsx 中注册 TbBuff")

# 4. 创建 buff.xlsx
buff_file = os.path.join(data_dir, 'buff.xlsx')
wb_buff = openpyxl.Workbook()
sheet = wb_buff.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'buff_type'
sheet.cell(1, 5).value = 'duration'
sheet.cell(1, 6).value = 'stack_limit'
sheet.cell(1, 7).value = 'refresh_on_apply'
sheet.cell(1, 8).value = 'effect'
sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'BuffType'
sheet.cell(2, 5).value = 'float#range=[0,3600]'
sheet.cell(2, 6).value = 'int#range=[1,99]'
sheet.cell(2, 7).value = 'bool'
sheet.cell(2, 8).value = 'BuffEffect'
sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 8).value = '$type'
sheet.cell(3, 9).value = 'stat'
sheet.cell(3, 10).value = 'value'
sheet.cell(3, 11).value = 'is_percent'
sheet.cell(3, 12).value = 'damage_per_tick'
sheet.cell(3, 13).value = 'tick_interval'
sheet.cell(3, 14).value = 'control_type'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据
sheet.append(['', 7001, 'buff_strength', 'POSITIVE|DISPELLABLE', 60.0, 5, True, 'StatModifier', 'atk', 20, False, '', '', ''])
sheet.append(['', 7002, 'buff_poison', 'NEGATIVE|STACKABLE', 30.0, 10, False, 'DotEffect', '', '', '', 10, 1.0, ''])
sheet.append(['', 7003, 'buff_stun', 'NEGATIVE', 5.0, 1, False, 'ControlEffect', '', '', '', '', '', 'stun'])

wb_buff.save(buff_file)
print("[OK] 已创建 buff.xlsx")

# ============================================================
# Task 3.3: 元素克制表 (TbElementRelation)
# ============================================================
print("\n[Task 3.3] 元素克制表 (TbElementRelation)")
print("-" * 60)

# 1. 添加 ElementRelation Bean
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active
last_row = sheet_beans.max_row

beans_data = [
    ['', 'ElementRelation', '', '', '', '', '元素克制关系'],
    ['', '', '', 'source_element', 'Element', '', '源元素'],
    ['', '', '', 'damage_multipliers', 'map,Element,float', '', '伤害倍率'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print("[OK] 已添加 ElementRelation Bean")

# 2. 创建 #ElementRelation.xlsx
elem_rel_file = os.path.join(data_dir, '#ElementRelation.xlsx')
wb_er = openpyxl.Workbook()
sheet = wb_er.active

sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'source_element'
sheet.cell(1, 3).value = '*damage_multipliers'
sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)

sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'Element'
sheet.cell(2, 3).value = 'map,Element,float'
sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)

sheet.cell(3, 1).value = '##var'
sheet.cell(3, 3).value = '$key'
sheet.cell(3, 4).value = 'value'

sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = '源元素'

# 数据: 火克制冰
row = 5
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 'FIRE'
sheet.cell(row, 3).value = 'ICE'
sheet.cell(row, 4).value = 1.5

row = 6
sheet.cell(row, 3).value = 'PHYSICAL'
sheet.cell(row, 4).value = 1.0

# 数据: 冰克制火
row = 7
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 'ICE'
sheet.cell(row, 3).value = 'FIRE'
sheet.cell(row, 4).value = 0.8

wb_er.save(elem_rel_file)
print("[OK] 已创建 #ElementRelation.xlsx")

print("\n" + "=" * 60)
print("Task 2.2 ~ Task 3.3 完成!")
print("=" * 60)
