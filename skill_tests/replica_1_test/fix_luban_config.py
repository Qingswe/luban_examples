import openpyxl
from openpyxl import load_workbook
import os
import shutil

# 修复Luban配置问题
# 问题：使用#前缀自动导入的表，其value_type也在__beans__.xlsx中定义，导致类型重复

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
datas_dir = os.path.join(replica_test_dir, 'Datas')
tables_file = os.path.join(datas_dir, '__tables__.xlsx')

print("=" * 80)
print("修复Luban配置问题")
print("=" * 80)

# 需要从#前缀改为普通文件名的表
tables_to_fix = [
    {
        'old_name': '#ElementRelation.xlsx',
        'new_name': 'element_relation.xlsx',
        'table_name': 'TbElementRelation',
        'value_type': 'ElementRelation'
    },
    {
        'old_name': '#Achievement.xlsx',
        'new_name': 'achievement.xlsx',
        'table_name': 'TbAchievement',
        'value_type': 'Achievement'
    },
    {
        'old_name': '#Dialog.xlsx',
        'new_name': 'dialog.xlsx',
        'table_name': 'TbDialog',
        'value_type': 'Dialog'
    },
    {
        'old_name': '#Recipe.xlsx',
        'new_name': 'recipe.xlsx',
        'table_name': 'TbRecipe',
        'value_type': 'Recipe'
    },
    {
        'old_name': '#Talent.xlsx',
        'new_name': 'talent.xlsx',
        'table_name': 'TbTalent',
        'value_type': 'TalentNode'
    },
    {
        'old_name': '#SkillEffect.xlsx',
        'new_name': 'skill_effect.xlsx',
        'table_name': 'TbSkillEffect',
        'value_type': 'SkillEffectConfig'
    },
    # #demo.item.xlsx特殊处理 - 它的value_type应该是Item，不需要修改
]

# 1. 重命名文件（移除#前缀）
print("\n步骤1: 重命名文件（移除#前缀）")
print("-" * 80)
for table_info in tables_to_fix:
    old_path = os.path.join(datas_dir, table_info['old_name'])
    new_path = os.path.join(datas_dir, table_info['new_name'])

    if os.path.exists(old_path):
        shutil.move(old_path, new_path)
        print(f"[OK] {table_info['old_name']} -> {table_info['new_name']}")
    else:
        print(f"[SKIP] {table_info['old_name']} 不存在")

# 2. 在__tables__.xlsx中添加这些表的注册
print("\n步骤2: 在__tables__.xlsx中注册表")
print("-" * 80)

wb = load_workbook(tables_file)
sheet = wb.active

# 找到最后一行
last_row = sheet.max_row

# 添加新表的注册
for table_info in tables_to_fix:
    last_row += 1
    sheet.cell(last_row, 1).value = ''  # ##var
    sheet.cell(last_row, 2).value = table_info['table_name']
    sheet.cell(last_row, 3).value = table_info['value_type']
    sheet.cell(last_row, 4).value = 'FALSE'  # read_schema_from_file
    sheet.cell(last_row, 5).value = table_info['new_name']
    sheet.cell(last_row, 6).value = ''  # index
    sheet.cell(last_row, 7).value = ''  # mode

    print(f"[OK] 已注册 {table_info['table_name']} -> {table_info['new_name']}")

wb.save(tables_file)
print(f"\n已更新 __tables__.xlsx")

print("\n" + "=" * 80)
print("修复完成！")
print("=" * 80)
print("\n已修复的问题:")
print("1. 移除了会导致类型重复的#前缀")
print("2. 在__tables__.xlsx中正确注册了所有表")
print("3. 设置read_schema_from_file=FALSE，从__beans__.xlsx读取schema")
