import openpyxl

wb = openpyxl.load_workbook('Datas/achievement.xlsx')
sheet = wb.active

print('achievement.xlsx当前结构:')
for row in range(1, min(9, sheet.max_row + 1)):
    print(f'Row {row}:')
    for col in range(1, 13):
        value = sheet.cell(row, col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        if value:
            print(f'  {col_letter}{row}: {repr(value)}')
    print()

print('合并单元格:')
for merged_cell in sheet.merged_cells.ranges:
    print(f'  {merged_cell}')
