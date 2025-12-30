import openpyxl
from openpyxl import Workbook
import os

# 创建最小化的equipment.xlsx，移除affix_pool字段

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
equipment_file = os.path.join(replica_test_dir, 'Datas', 'equipment.xlsx')

print("创建最小化equipment.xlsx（暂时移除affix_pool）...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'slot'
sheet.cell(1, 5).value = 'base_stats'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
sheet.cell(1, 8).value = 'set_id'

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'EquipSlot'
sheet.cell(2, 5).value = 'map,string,int'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
sheet.cell(2, 8).value = 'int?'

# 第3行: ##var (子字段)
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 5).value = 'hp'
sheet.cell(3, 6).value = 'atk'
sheet.cell(3, 7).value = 'def'

# 第4行: ##
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'

# 数据行1
row = 5
sheet.cell(row, 2).value = 4001
sheet.cell(row, 3).value = 'Iron Sword'
sheet.cell(row, 4).value = 'WEAPON'
sheet.cell(row, 5).value = 0
sheet.cell(row, 6).value = 30
sheet.cell(row, 7).value = 5
sheet.cell(row, 8).value = ''

# 数据行2
row = 6
sheet.cell(row, 2).value = 4002
sheet.cell(row, 3).value = 'Steel Armor'
sheet.cell(row, 4).value = 'ARMOR'
sheet.cell(row, 5).value = 100
sheet.cell(row, 6).value = 0
sheet.cell(row, 7).value = 50
sheet.cell(row, 8).value = 1001

wb.save(equipment_file)
print(f"[OK] 已创建最小化 equipment.xlsx")
print("  - 暂时移除affix_pool字段以定位问题")

print("\n同时需要更新__beans__.xlsx中Equipment的定义，移除affix_pool字段")
