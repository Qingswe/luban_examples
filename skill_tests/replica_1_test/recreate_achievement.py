import openpyxl
from openpyxl import Workbook
import os

# 完全重新创建achievement.xlsx，确保合并单元格正确

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
file_path = os.path.join(replica_test_dir, 'Datas', 'achievement.xlsx')

print("完全重新创建achievement.xlsx...")

wb = Workbook()
sheet = wb.active

# 第1行：##var行
# 先设置单元格值，再合并
sheet['A1'] = '##var'
sheet['B1'] = 'id'
sheet['C1'] = 'name'
sheet['D1'] = 'desc'
sheet['E1'] = 'condition'  # 将被合并到E1:H1
sheet['I1'] = '*rewards'    # 将被合并到I1:K1
sheet['L1'] = 'points'

# 第2行：##type行
sheet['A2'] = '##type'
sheet['B2'] = 'int'
sheet['C2'] = 'text'
sheet['D2'] = 'text'
sheet['E2'] = 'AchievementCondition'  # 将被合并到E2:H2
sheet['I2'] = 'list,AchievementReward'  # 将被合并到I2:K2
sheet['L2'] = 'int'

# 执行合并（在设置值之后）
sheet.merge_cells('E1:H1')  # condition字段
sheet.merge_cells('E2:H2')  # AchievementCondition类型
sheet.merge_cells('I1:K1')  # *rewards字段
sheet.merge_cells('I2:K2')  # list,AchievementReward类型

# 第3行：##var子字段行
sheet['A3'] = '##var'
# condition子字段
sheet['E3'] = '$type'
sheet['F3'] = 'monster_type'
sheet['G3'] = 'count'
sheet['H3'] = 'item_id'
# *rewards的元素字段（AchievementReward）
sheet['J3'] = 'item_id'
sheet['K3'] = 'count'

# 第4行：##注释行
sheet['A4'] = '##'
sheet['B4'] = 'ID'
sheet['C4'] = '名称'

# 数据行1：KillCount成就
sheet['B5'] = 10001
sheet['C5'] = 'ach_kill_goblin'
sheet['D5'] = 'Kill 100 goblins'
sheet['E5'] = 'KillCount'
sheet['F5'] = 'goblin'
sheet['G5'] = 100
sheet['J5'] = 1001  # reward item_id
sheet['K5'] = 10    # reward count
sheet['L5'] = 50    # points

# 数据行2：第二个reward（属于成就10001）
sheet['J6'] = 1002
sheet['K6'] = 5

# 数据行3：CollectItem成就
sheet['B7'] = 10002
sheet['C7'] = 'ach_collect_gold'
sheet['D7'] = 'Collect 50 gold'
sheet['E7'] = 'CollectItem'
sheet['G7'] = 50
sheet['H7'] = 2002
sheet['J7'] = 3001
sheet['K7'] = 1
sheet['L7'] = 100

wb.save(file_path)
print(f"[OK] 已完全重新创建 achievement.xlsx")
print("  - condition字段合并E1:H1, E2:H2")
print("  - *rewards字段合并I1:K1, I2:K2")
print("  - 子字段定义在第3行")
