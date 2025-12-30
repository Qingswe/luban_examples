import openpyxl
from openpyxl import load_workbook
import os

# 更新skill_tree.xlsx中的$type列

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
skill_tree_file = os.path.join(replica_test_dir, 'Datas', 'skill_tree.xlsx')

print("更新skill_tree.xlsx中的$type列...")

wb = load_workbook(skill_tree_file)
sheet = wb.active

# 找到$type列（应该在第6列）
type_col = None
for col_idx in range(1, sheet.max_column + 1):
    if sheet.cell(3, col_idx).value == '$type':
        type_col = col_idx
        break

if not type_col:
    print("[ERROR] 未找到$type列")
    exit(1)

print(f"找到$type列在第{type_col}列")

# 从第5行开始更新数据
update_count = 0
for row_idx in range(5, sheet.max_row + 1):
    type_value = sheet.cell(row_idx, type_col).value

    if type_value == 'DamageEffect':
        sheet.cell(row_idx, type_col).value = 'SkillDamageEffect'
        print(f"  Row {row_idx}: DamageEffect -> SkillDamageEffect")
        update_count += 1
    elif type_value == 'HealEffect':
        sheet.cell(row_idx, type_col).value = 'SkillHealEffect'
        print(f"  Row {row_idx}: HealEffect -> SkillHealEffect")
        update_count += 1
    elif type_value == 'BuffEffect':
        sheet.cell(row_idx, type_col).value = 'SkillBuffEffect'
        print(f"  Row {row_idx}: BuffEffect -> SkillBuffEffect")
        update_count += 1

wb.save(skill_tree_file)
print(f"\n[OK] 已更新 {update_count} 行数据")
