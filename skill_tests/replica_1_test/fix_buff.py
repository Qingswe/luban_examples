import openpyxl
from openpyxl import Workbook
import os

# 修复buff.xlsx - 正确分离BuffEffect子类字段

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
buff_file = os.path.join(replica_test_dir, 'Datas', 'buff.xlsx')

print("重新创建buff.xlsx，正确处理BuffEffect子类字段...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var - Buff字段 + effect (BuffEffect)
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'buff_type'
sheet.cell(1, 5).value = 'duration'
sheet.cell(1, 6).value = 'stack_limit'
sheet.cell(1, 7).value = 'refresh_on_apply'
sheet.cell(1, 8).value = 'effect'
sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'BuffType'
sheet.cell(2, 5).value = 'float#range=[0,3600]'
sheet.cell(2, 6).value = 'int#range=[1,99]'
sheet.cell(2, 7).value = 'bool'
sheet.cell(2, 8).value = 'BuffEffect'
sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)

# 第3行: ##var - BuffEffect子类的字段
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 8).value = '$type'
# StatModifier字段
sheet.cell(3, 9).value = 'stat'
sheet.cell(3, 10).value = 'value'
sheet.cell(3, 11).value = 'is_percent'
# DotEffect字段 (重用列)
# sheet.cell(3, 9).value = 'damage_per_tick'  # 与stat冲突，需要分离
# sheet.cell(3, 10).value = 'tick_interval'
# ControlEffect字段
# sheet.cell(3, 9).value = 'control_type'

# 注意：由于不同子类字段位置不同，我们需要更复杂的结构
# 让我使用更简单的方法：每个子类的字段单独命名

sheet.cell(3, 9).value = 'stat'  # StatModifier
sheet.cell(3, 10).value = 'value'  # StatModifier
sheet.cell(3, 11).value = 'is_percent'  # StatModifier
sheet.cell(3, 12).value = 'damage_per_tick'  # DotEffect
sheet.cell(3, 13).value = 'tick_interval'  # DotEffect
sheet.cell(3, 14).value = 'control_type'  # ControlEffect

# 第4行: ##
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据行1: StatModifier buff
row = 5
sheet.cell(row, 2).value = 7001
sheet.cell(row, 3).value = 'buff_strength'
sheet.cell(row, 4).value = 'POSITIVE|DISPELLABLE'
sheet.cell(row, 5).value = 60
sheet.cell(row, 6).value = 5
sheet.cell(row, 7).value = 'True'
sheet.cell(row, 8).value = 'StatModifier'
sheet.cell(row, 9).value = 'atk'
sheet.cell(row, 10).value = 20
sheet.cell(row, 11).value = 'False'

# 数据行2: DotEffect buff
row = 6
sheet.cell(row, 2).value = 7002
sheet.cell(row, 3).value = 'buff_poison'
sheet.cell(row, 4).value = 'NEGATIVE|STACKABLE'
sheet.cell(row, 5).value = 30
sheet.cell(row, 6).value = 10
sheet.cell(row, 7).value = 'False'
sheet.cell(row, 8).value = 'DotEffect'
sheet.cell(row, 12).value = 10
sheet.cell(row, 13).value = 1

# 数据行3: ControlEffect buff
row = 7
sheet.cell(row, 2).value = 7003
sheet.cell(row, 3).value = 'buff_stun'
sheet.cell(row, 4).value = 'NEGATIVE'
sheet.cell(row, 5).value = 5
sheet.cell(row, 6).value = 1
sheet.cell(row, 7).value = 'False'
sheet.cell(row, 8).value = 'ControlEffect'
sheet.cell(row, 14).value = 'stun'

wb.save(buff_file)
print(f"[OK] 已重新创建 buff.xlsx")
print("  - 正确分离了StatModifier、DotEffect、ControlEffect的字段")
print("  - 每个子类的字段在不同的列位置")
