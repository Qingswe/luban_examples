import openpyxl
from openpyxl import load_workbook, Workbook
import os

# 重新创建equipment.xlsx，确保格式正确

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
equipment_file = os.path.join(replica_test_dir, 'Datas', 'equipment.xlsx')

print("重新创建equipment.xlsx...")

# 读取原有数据
wb_old = load_workbook(equipment_file)
sheet_old = wb_old.active

# 提取数据行（从第5行开始）
data_rows = []
for row_idx in range(5, sheet_old.max_row + 1):
    row_data = []
    for col_idx in range(1, 15):
        row_data.append(sheet_old.cell(row_idx, col_idx).value)
    data_rows.append(row_data)

# 创建新文件
wb = Workbook()
sheet = wb.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'slot'
sheet.cell(1, 5).value = 'base_stats'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
sheet.cell(1, 8).value = 'affix_pool'
sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
sheet.cell(1, 14).value = 'set_id'

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'EquipSlot'
sheet.cell(2, 5).value = 'map,string,int'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
sheet.cell(2, 8).value = '(list#index=id),Affix'
sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)
sheet.cell(2, 14).value = 'int?'

# 第3行: ##var (子字段)
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 5).value = 'hp'
sheet.cell(3, 6).value = 'atk'
sheet.cell(3, 7).value = 'def'
sheet.cell(3, 8).value = 'id'
sheet.cell(3, 9).value = 'name'
sheet.cell(3, 10).value = '$type'
sheet.cell(3, 11).value = 'stat'
sheet.cell(3, 12).value = 'value'
sheet.cell(3, 13).value = 'percent'

# 第4行: ##
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'
sheet.cell(4, 4).value = '槽位'
sheet.cell(4, 5).value = 'HP'
sheet.cell(4, 6).value = 'ATK'
sheet.cell(4, 7).value = 'DEF'

# 写入数据行
for row_data in data_rows:
    sheet.append(row_data)

wb.save(equipment_file)
print(f"[OK] 已重新创建 equipment.xlsx")
print(f"  - 已正确设置标题头和单元格合并")
print(f"  - 已保留原有数据 ({len(data_rows)} 行)")
