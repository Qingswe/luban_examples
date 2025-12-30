import openpyxl
from openpyxl import Workbook
import os

# 正确创建equipment.xlsx
# 问题：affix_pool字段需要正确的列结构

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
equipment_file = os.path.join(replica_test_dir, 'Datas', 'equipment.xlsx')

print("正确创建equipment.xlsx...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'slot'
sheet.cell(1, 5).value = 'base_stats'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
sheet.cell(1, 8).value = 'affix_pool'
sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=14)
sheet.cell(1, 15).value = 'set_id'

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'EquipSlot'
sheet.cell(2, 5).value = 'map,string,int'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
sheet.cell(2, 8).value = '(list#index=id),Affix'
sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=14)
sheet.cell(2, 15).value = 'int?'

# 第3行: ##var (子字段 - affix_pool的结构)
sheet.cell(3, 1).value = '##var'
# base_stats的子字段
sheet.cell(3, 5).value = 'hp'
sheet.cell(3, 6).value = 'atk'
sheet.cell(3, 7).value = 'def'
# affix_pool (Affix) 的字段
sheet.cell(3, 8).value = 'id'
sheet.cell(3, 9).value = 'name'
sheet.cell(3, 10).value = 'modifier'
sheet.merge_cells(start_row=3, start_column=10, end_row=3, end_column=14)

# 第4行: ##var (Modifier的子字段 - 第二层嵌套)
sheet.cell(4, 1).value = '##var'
sheet.cell(4, 10).value = '$type'
sheet.cell(4, 11).value = 'stat'
sheet.cell(4, 12).value = 'value'
sheet.cell(4, 13).value = 'percent'

# 第5行: ##
sheet.cell(5, 1).value = '##'
sheet.cell(5, 2).value = 'ID'
sheet.cell(5, 3).value = '名称'
sheet.cell(5, 4).value = '槽位'

# 数据行1: 铁剑
row = 6
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 4001
sheet.cell(row, 3).value = 'Iron Sword'
sheet.cell(row, 4).value = 'WEAPON'
sheet.cell(row, 5).value = 0      # hp
sheet.cell(row, 6).value = 30     # atk
sheet.cell(row, 7).value = 5      # def
sheet.cell(row, 8).value = 101    # affix id
sheet.cell(row, 9).value = 'Sharp'  # affix name
sheet.cell(row, 10).value = 'FlatModifier'  # modifier $type
sheet.cell(row, 11).value = 'atk'  # modifier stat
sheet.cell(row, 12).value = 10     # modifier value
sheet.cell(row, 13).value = ''     # modifier percent
sheet.cell(row, 15).value = ''     # set_id

# 数据行2: 钢铠甲
row = 7
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 4002
sheet.cell(row, 3).value = 'Steel Armor'
sheet.cell(row, 4).value = 'ARMOR'
sheet.cell(row, 5).value = 100    # hp
sheet.cell(row, 6).value = 0      # atk
sheet.cell(row, 7).value = 50     # def
sheet.cell(row, 8).value = 102    # affix id
sheet.cell(row, 9).value = 'Tough'  # affix name
sheet.cell(row, 10).value = 'FlatModifier'  # modifier $type
sheet.cell(row, 11).value = 'def'  # modifier stat
sheet.cell(row, 12).value = 20     # modifier value
sheet.cell(row, 13).value = ''     # modifier percent
sheet.cell(row, 15).value = 1001   # set_id

wb.save(equipment_file)
print(f"[OK] 已正确创建 equipment.xlsx")
print("  - 使用正确的嵌套结构: Equipment -> Affix -> Modifier")
print("  - modifier字段有独立的##var行定义子字段")
