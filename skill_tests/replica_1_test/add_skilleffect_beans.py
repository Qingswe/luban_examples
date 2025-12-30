import openpyxl
from openpyxl import load_workbook
import os

# 添加SkillEffect多态类型到__beans__.xlsx

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
beans_file = os.path.join(replica_test_dir, 'Datas', '__beans__.xlsx')

print("添加SkillEffect多态类型到__beans__.xlsx...")

wb = load_workbook(beans_file)
sheet = wb.active

# 找到最后一行
last_row = sheet.max_row + 1

# 添加SkillEffect多态层次（Task 1.2的需求）
beans_data = [
    ('SkillEffect', '', 'SkillEffect基类', []),

    ('DamageEffect', 'SkillEffect', '伤害效果', [
        ('damage', 'int', '', '伤害值'),
        ('element', 'string', '', '元素类型'),
    ]),

    ('HealEffect', 'SkillEffect', '治疗效果', [
        ('heal_amount', 'int', '', '治疗量'),
        ('heal_percent', 'float', '', '治疗百分比'),
    ]),

    ('BuffEffect', 'SkillEffect', 'Buff效果', [
        ('buff_id', 'int', '', 'BuffID'),
        ('duration', 'float', '', '持续时间'),
    ]),
]

for full_name, parent, comment, fields in beans_data:
    # Bean定义行
    sheet.cell(last_row, 1).value = ''  # ##var列为空
    sheet.cell(last_row, 2).value = full_name
    sheet.cell(last_row, 3).value = parent
    sheet.cell(last_row, 7).value = comment
    last_row += 1

    # 字段定义行
    for field_name, field_type, field_group, field_comment in fields:
        sheet.cell(last_row, 1).value = ''
        sheet.cell(last_row, 2).value = ''
        sheet.cell(last_row, 3).value = ''
        sheet.cell(last_row, 10).value = field_name
        sheet.cell(last_row, 11).value = ''  # alias
        sheet.cell(last_row, 12).value = field_type
        sheet.cell(last_row, 13).value = field_group
        sheet.cell(last_row, 14).value = field_comment
        sheet.cell(last_row, 15).value = ''  # tags
        sheet.cell(last_row, 16).value = ''  # variants
        last_row += 1

wb.save(beans_file)
print(f"[OK] 已添加 SkillEffect 多态类型")
print("     - SkillEffect (基类)")
print("     - DamageEffect (伤害效果)")
print("     - HealEffect (治疗效果)")
print("     - BuffEffect (Buff效果)")
