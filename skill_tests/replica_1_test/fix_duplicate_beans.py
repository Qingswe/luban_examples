import openpyxl
from openpyxl import load_workbook
import os

# 修复DamageEffect重复问题
# 将Task 1.2 SkillEffect层次中的类型重命名以避免与Task 3.1 Effect层次冲突

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
beans_file = os.path.join(replica_test_dir, 'Datas', '__beans__.xlsx')

print("修复DamageEffect重复问题...")
print("将Task 1.2的SkillEffect子类重命名以避免冲突")

wb = load_workbook(beans_file)
sheet = wb.active

# 查找并重命名最后添加的SkillEffect层次中的类型
# 将 DamageEffect -> SkillDamageEffect
# 将 HealEffect -> SkillHealEffect
# 将 BuffEffect -> SkillBuffEffect

renamed_count = 0
for row_idx in range(sheet.max_row, 0, -1):
    full_name = sheet.cell(row_idx, 2).value
    parent = sheet.cell(row_idx, 3).value

    # 找到SkillEffect的子类（倒序查找最近添加的）
    if parent == 'SkillEffect':
        if full_name == 'DamageEffect':
            sheet.cell(row_idx, 2).value = 'SkillDamageEffect'
            print(f"  Row {row_idx}: DamageEffect -> SkillDamageEffect")
            renamed_count += 1
        elif full_name == 'HealEffect':
            sheet.cell(row_idx, 2).value = 'SkillHealEffect'
            print(f"  Row {row_idx}: HealEffect -> SkillHealEffect")
            renamed_count += 1
        elif full_name == 'BuffEffect':
            sheet.cell(row_idx, 2).value = 'SkillBuffEffect'
            print(f"  Row {row_idx}: BuffEffect -> SkillBuffEffect")
            renamed_count += 1

        if renamed_count >= 3:
            break

wb.save(beans_file)
print(f"\n[OK] 已重命名 {renamed_count} 个Bean类型")
print("\n注意: skill_tree.xlsx中的$type列也需要相应更新!")
