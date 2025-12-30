import openpyxl
from openpyxl import Workbook
import os

# 重新创建element_relation.xlsx，使用正确的map格式（列限定格式）

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
file_path = os.path.join(replica_test_dir, 'Datas', 'element_relation.xlsx')

print("重新创建element_relation.xlsx...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'source_element'
sheet.cell(1, 3).value = 'damage_multipliers'
sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'Element'
sheet.cell(2, 3).value = 'map,Element,float'
sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=6)

# 第3行: ##var (所有Element作为key)
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 3).value = 'FIRE'
sheet.cell(3, 4).value = 'ICE'
sheet.cell(3, 5).value = 'LIGHTNING'
sheet.cell(3, 6).value = 'PHYSICAL'

# 第4行: ##
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = '源元素'

# 数据行1: FIRE的伤害倍率
row = 5
sheet.cell(row, 2).value = 'FIRE'
sheet.cell(row, 3).value = 1.0   # FIRE -> FIRE = 1.0
sheet.cell(row, 4).value = 1.5   # FIRE -> ICE = 1.5
sheet.cell(row, 5).value = 1.0   # FIRE -> LIGHTNING = 1.0
sheet.cell(row, 6).value = 1.0   # FIRE -> PHYSICAL = 1.0

# 数据行2: ICE的伤害倍率
row = 6
sheet.cell(row, 2).value = 'ICE'
sheet.cell(row, 3).value = 0.8   # ICE -> FIRE = 0.8
sheet.cell(row, 4).value = 1.0   # ICE -> ICE = 1.0
sheet.cell(row, 5).value = 1.2   # ICE -> LIGHTNING = 1.2
sheet.cell(row, 6).value = 1.0   # ICE -> PHYSICAL = 1.0

# 数据行3: LIGHTNING的伤害倍率
row = 7
sheet.cell(row, 2).value = 'LIGHTNING'
sheet.cell(row, 3).value = 1.0   # LIGHTNING -> FIRE = 1.0
sheet.cell(row, 4).value = 1.3   # LIGHTNING -> ICE = 1.3
sheet.cell(row, 5).value = 1.0   # LIGHTNING -> LIGHTNING = 1.0
sheet.cell(row, 6).value = 1.5   # LIGHTNING -> PHYSICAL = 1.5

# 数据行4: PHYSICAL的伤害倍率
row = 8
sheet.cell(row, 2).value = 'PHYSICAL'
sheet.cell(row, 3).value = 1.0   # PHYSICAL -> FIRE = 1.0
sheet.cell(row, 4).value = 1.0   # PHYSICAL -> ICE = 1.0
sheet.cell(row, 5).value = 0.7   # PHYSICAL -> LIGHTNING = 0.7
sheet.cell(row, 6).value = 1.0   # PHYSICAL -> PHYSICAL = 1.0

wb.save(file_path)
print(f"[OK] 已重新创建 element_relation.xlsx")
print("  - 使用列限定格式，每个Element作为key列")
print("  - 每行定义一个源元素对所有目标元素的伤害倍率")
