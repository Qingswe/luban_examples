import openpyxl
from openpyxl import Workbook
import os

# 简化equipment.xlsx，移除index验证器来测试

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
equipment_file = os.path.join(replica_test_dir, 'Datas', 'equipment.xlsx')

print("创建简化版equipment.xlsx（移除index验证器）...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'slot'
sheet.cell(1, 5).value = 'base_stats'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
sheet.cell(1, 8).value = 'affix_pool'
sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=14)
sheet.cell(1, 15).value = 'set_id'

# 第2行: ##type - 移除index验证器
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'EquipSlot'
sheet.cell(2, 5).value = 'map,string,int'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
sheet.cell(2, 8).value = 'list,Affix'  # 简化：移除index
sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=14)
sheet.cell(2, 15).value = 'int?'

# 第3行: ##var (子字段)
sheet.cell(3, 1).value = '##var'
# base_stats的子字段
sheet.cell(3, 5).value = 'hp'
sheet.cell(3, 6).value = 'atk'
sheet.cell(3, 7).value = 'def'
# affix_pool (Affix) 的字段 - 按照__beans__.xlsx中定义的顺序
sheet.cell(3, 8).value = 'id'
sheet.cell(3, 9).value = 'name'
sheet.cell(3, 10).value = 'modifier'
sheet.merge_cells(start_row=3, start_column=10, end_row=3, end_column=14)

# 第4行: ##var (Modifier的子字段)
sheet.cell(4, 1).value = '##var'
sheet.cell(4, 10).value = '$type'
sheet.cell(4, 11).value = 'stat'
sheet.cell(4, 12).value = 'value'
sheet.cell(4, 13).value = 'percent'

# 第5行: ##
sheet.cell(5, 1).value = '##'
sheet.cell(5, 2).value = 'ID'

# 数据行1
row = 6
sheet.cell(row, 2).value = 4001
sheet.cell(row, 3).value = 'Iron Sword'
sheet.cell(row, 4).value = 'WEAPON'
sheet.cell(row, 5).value = 0
sheet.cell(row, 6).value = 30
sheet.cell(row, 7).value = 5
sheet.cell(row, 8).value = 101
sheet.cell(row, 9).value = 'Sharp'
sheet.cell(row, 10).value = 'FlatModifier'
sheet.cell(row, 11).value = 'atk'
sheet.cell(row, 12).value = 10
sheet.cell(row, 15).value = ''

wb.save(equipment_file)
print(f"[OK] 已创建简化版 equipment.xlsx")
print("  - 移除了index验证器")
print("  - 保持了正确的嵌套结构")
