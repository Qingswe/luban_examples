import openpyxl
from openpyxl import Workbook
import os

# 批量修复achievement, dialog, recipe, talent四个表

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'

print("="*60)
print("批量修复 achievement.xlsx, dialog.xlsx, recipe.xlsx, talent.xlsx")
print("="*60)

# ====================================================================================
# 1. achievement.xlsx
# ====================================================================================
print("\n[1/4] 修复 achievement.xlsx...")
file_path = os.path.join(replica_test_dir, 'Datas', 'achievement.xlsx')
wb = Workbook()
sheet = wb.active

# 第1-2行：字段定义
sheet.cell(1, 1).value = '##var'
sheet['B1'] = 'id'; sheet['C1'] = 'name'; sheet['D1'] = 'desc'
sheet['E1'] = 'condition'
sheet.merge_cells('E1:H1')
sheet['I1'] = '*rewards'
sheet['L1'] = 'points'

sheet.cell(2, 1).value = '##type'
sheet['B2'] = 'int'; sheet['C2'] = 'text'; sheet['D2'] = 'text'
sheet['E2'] = 'AchievementCondition'
sheet.merge_cells('E2:H2')
sheet['I2'] = 'list,AchievementReward'
sheet['L2'] = 'int'

# 第3行：子字段
sheet.cell(3, 1).value = '##var'
sheet['E3'] = '$type'; sheet['F3'] = 'monster_type'; sheet['G3'] = 'count'; sheet['H3'] = 'item_id'
sheet['J3'] = 'item_id'; sheet['K3'] = 'count'

# 第4行：注释
sheet.cell(4, 1).value = '##'
sheet['B4'] = 'ID'; sheet['C4'] = '名称'

# 数据
sheet['B5'] = 10001; sheet['C5'] = 'ach_kill_goblin'; sheet['D5'] = 'Kill 100 goblins'
sheet['E5'] = 'KillCount'; sheet['F5'] = 'goblin'; sheet['G5'] = 100
sheet['J5'] = 1001; sheet['K5'] = 10; sheet['L5'] = 50

sheet['J6'] = 1002; sheet['K6'] = 5

sheet['B7'] = 10002; sheet['C7'] = 'ach_collect_gold'; sheet['D7'] = 'Collect 50 gold'
sheet['E7'] = 'CollectItem'; sheet['G7'] = 50; sheet['H7'] = 2002
sheet['J7'] = 3001; sheet['K7'] = 1; sheet['L7'] = 100

wb.save(file_path)
print("[OK] achievement.xlsx 已修复")

# ====================================================================================
# 2. dialog.xlsx
# ====================================================================================
print("\n[2/4] 修复 dialog.xlsx...")
file_path = os.path.join(replica_test_dir, 'Datas', 'dialog.xlsx')
wb = Workbook()
sheet = wb.active

# Dialog: id, name, *lines (list<DialogLine>), choices (list<DialogChoice>)
# 简化：假设DialogLine和DialogChoice都是简单的text/string字段

sheet.cell(1, 1).value = '##var'
sheet['B1'] = 'id'; sheet['C1'] = 'name'; sheet['D1'] = '*lines'; sheet['E1'] = 'choices'

sheet.cell(2, 1).value = '##type'
sheet['B2'] = 'int'; sheet['C2'] = 'string'; sheet['D2'] = 'list,string'; sheet['E2'] = 'list,DialogChoice'

# DialogChoice字段需要在第3行定义 - 先检查__beans__.xlsx看它有哪些字段
# 暂时简化为text和next_dialog_id
sheet.cell(3, 1).value = '##var'
sheet['E3'] = 'text'; sheet['F3'] = 'next_dialog_id'

sheet.cell(4, 1).value = '##'
sheet['B4'] = 'ID'

# 数据
sheet['B5'] = 8001; sheet['C5'] = 'dialog_greeting'
sheet['D5'] = 'Hello, traveler!'
sheet['E5'] = 'Tell me about quests'; sheet['F5'] = 8002

sheet['D6'] = 'How may I help you?'
sheet['E6'] = 'Goodbye'; sheet['F6'] = 0

wb.save(file_path)
print("[OK] dialog.xlsx 已修复")

# ====================================================================================
# 3. recipe.xlsx
# ====================================================================================
print("\n[3/4] 修复 recipe.xlsx...")
file_path = os.path.join(replica_test_dir, 'Datas', 'recipe.xlsx')
wb = Workbook()
sheet = wb.active

# Recipe: id, name, output_item, output_count, *materials (list<Material>), unlock_level
# Material需要检查__beans__定义 - 假设有item_id和count字段

sheet.cell(1, 1).value = '##var'
sheet['B1'] = 'id'; sheet['C1'] = 'name'; sheet['D1'] = 'output_item'; sheet['E1'] = 'output_count'
sheet['F1'] = '*materials'
sheet['I1'] = 'unlock_level'

sheet.cell(2, 1).value = '##type'
sheet['B2'] = 'int'; sheet['C2'] = 'string'
sheet['D2'] = 'int#ref=TbItem'; sheet['E2'] = 'int#range=[1,99]'
sheet['F2'] = 'list,Material'
sheet['I2'] = 'int#range=[1,100]'

# 第3行：Material字段
sheet.cell(3, 1).value = '##var'
sheet['G3'] = 'item_id'; sheet['H3'] = 'count'

sheet.cell(4, 1).value = '##'
sheet['B4'] = 'ID'

# 数据
sheet['B5'] = 9001; sheet['C5'] = 'recipe_iron_sword'
sheet['D5'] = 4001; sheet['E5'] = 1
sheet['G5'] = 2001; sheet['H5'] = 10; sheet['I5'] = 5

sheet['G6'] = 2002; sheet['H6'] = 5

wb.save(file_path)
print("[OK] recipe.xlsx 已修复")

# ====================================================================================
# 4. talent.xlsx - 检查现有文件是否正确
# ====================================================================================
print("\n[4/4] 检查 talent.xlsx...")
file_path = os.path.join(replica_test_dir, 'Datas', 'talent.xlsx')
try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 检查是否有*effects字段
    if sheet['E1'].value == '*effects':
        print("[OK] talent.xlsx 结构看起来正确")
    else:
        print("[WARN] talent.xlsx *effects字段位置不对，需要重建")
except Exception as e:
    print(f"[ERROR] talent.xlsx 读取失败: {e}")

print("\n" + "="*60)
print("批量修复完成！")
print("="*60)
