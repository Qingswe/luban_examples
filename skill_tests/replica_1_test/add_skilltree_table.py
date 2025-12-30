import openpyxl
from openpyxl import load_workbook
import os

# 添加TbSkillTree的注册

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
tables_file = os.path.join(replica_test_dir, 'Datas', '__tables__.xlsx')

print("添加TbSkillTree到__tables__.xlsx...")

wb = load_workbook(tables_file)
sheet = wb.active

# 找到最后一行
last_row = sheet.max_row + 1

# 添加TbSkillTree注册
# 根据Task 1.2，skill_tree.xlsx应该是手动注册，read_schema_from_file=TRUE
sheet.cell(last_row, 1).value = ''  # ##var
sheet.cell(last_row, 2).value = 'TbSkillTree'
sheet.cell(last_row, 3).value = 'SkillNode'
sheet.cell(last_row, 4).value = 'TRUE'  # read_schema_from_file - 从Excel读取schema
sheet.cell(last_row, 5).value = 'skill_tree.xlsx'
sheet.cell(last_row, 6).value = ''  # index
sheet.cell(last_row, 7).value = ''  # mode

wb.save(tables_file)
print(f"[OK] 已注册 TbSkillTree -> skill_tree.xlsx")
print(f"     value_type: SkillNode")
print(f"     read_schema_from_file: TRUE (从Excel标题头读取schema)")
