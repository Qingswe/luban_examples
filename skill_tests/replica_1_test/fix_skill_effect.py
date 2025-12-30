import openpyxl
from openpyxl import Workbook
import os

# 重新创建skill_effect.xlsx - 修正字段列分配

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
file_path = os.path.join(replica_test_dir, 'Datas', 'skill_effect.xlsx')

print("重新创建skill_effect.xlsx...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var
# SkillEffectConfig字段: id, name, effect, scaling_stat, scaling_factor
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'effect'
sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=10)
sheet.cell(1, 11).value = 'scaling_stat'
sheet.cell(1, 12).value = 'scaling_factor'

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'string'
sheet.cell(2, 4).value = 'Effect'
sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=10)
sheet.cell(2, 11).value = 'string'
sheet.cell(2, 12).value = 'float'

# 第3行: ##var - Effect子类字段
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 4).value = '$type'
# PhysicalDamage & MagicalDamage共享
sheet.cell(3, 5).value = 'base_damage'
sheet.cell(3, 6).value = 'armor_pen'      # PhysicalDamage
sheet.cell(3, 7).value = 'element'        # MagicalDamage
sheet.cell(3, 8).value = 'ignore_resist'  # MagicalDamage
# HealEffect
sheet.cell(3, 9).value = 'heal_base'
sheet.cell(3, 10).value = 'heal_scale'

# 第4行: ##
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据行1: PhysicalDamage
row = 5
sheet.cell(row, 2).value = 6001
sheet.cell(row, 3).value = 'Physical Strike'
sheet.cell(row, 4).value = 'PhysicalDamage'
sheet.cell(row, 5).value = 100          # base_damage
sheet.cell(row, 6).value = 0.3          # armor_pen
sheet.cell(row, 11).value = 'atk'       # scaling_stat
sheet.cell(row, 12).value = 1.5         # scaling_factor

# 数据行2: MagicalDamage
row = 6
sheet.cell(row, 2).value = 6002
sheet.cell(row, 3).value = 'Fire Blast'
sheet.cell(row, 4).value = 'MagicalDamage'
sheet.cell(row, 5).value = 150          # base_damage
sheet.cell(row, 7).value = 'FIRE'       # element
sheet.cell(row, 8).value = 'True'       # ignore_resist
sheet.cell(row, 11).value = 'mp'        # scaling_stat
sheet.cell(row, 12).value = 2.0         # scaling_factor

# 数据行3: HealEffect
row = 7
sheet.cell(row, 2).value = 6003
sheet.cell(row, 3).value = 'Heal'
sheet.cell(row, 4).value = 'HealEffect'
sheet.cell(row, 9).value = 80           # heal_base
sheet.cell(row, 10).value = 1.2         # heal_scale
sheet.cell(row, 11).value = 'mp'        # scaling_stat
sheet.cell(row, 12).value = 1.0         # scaling_factor

wb.save(file_path)
print(f"[OK] 已重新创建 skill_effect.xlsx")
print("  - effect字段合并列4-10 (包含所有Effect子类字段)")
print("  - scaling_stat和scaling_factor在列11-12")
