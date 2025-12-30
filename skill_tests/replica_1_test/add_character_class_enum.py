import openpyxl
from openpyxl import load_workbook
import os

# 添加缺失的CharacterClass枚举

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
enums_file = os.path.join(replica_test_dir, 'Datas', '__enums__.xlsx')

print("添加CharacterClass枚举...")

wb = load_workbook(enums_file)
sheet = wb.active

# 找到最后一行
last_row = sheet.max_row

# 添加CharacterClass枚举
enums_data = [
    ['', 'CharacterClass', 'FALSE', 'TRUE', '', '角色职业', '', 'WARRIOR', '战士', '1', '战士职业'],
    ['', '', '', '', '', '', '', 'MAGE', '法师', '2', '法师职业'],
    ['', '', '', '', '', '', '', 'ARCHER', '弓手', '3', '弓手职业'],
    ['', '', '', '', '', '', '', 'PRIEST', '牧师', '4', '牧师职业'],
]

for row_data in enums_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet.cell(last_row, col_idx).value = value

wb.save(enums_file)
print(f"[OK] 已添加 CharacterClass 枚举到 __enums__.xlsx")
print(f"  - WARRIOR (战士) = 1")
print(f"  - MAGE (法师) = 2")
print(f"  - ARCHER (弓手) = 3")
print(f"  - PRIEST (牧师) = 4")
