import openpyxl
from openpyxl import Workbook
import os

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'

# ====================================================================================
# 1. dialog.xlsx
# ====================================================================================
print("[1/2] 重新创建dialog.xlsx...")
file_path = os.path.join(replica_test_dir, 'Datas', 'dialog.xlsx')

wb = Workbook()
sheet = wb.active

# 第1行
sheet['A1'] = '##var'
sheet['B1'] = 'id'
sheet['C1'] = 'name'
sheet['D1'] = '*lines'
sheet['E1'] = 'choices'

# 第2行
sheet['A2'] = '##type'
sheet['B2'] = 'int'
sheet['C2'] = 'string'
sheet['D2'] = 'list,string'
sheet['E2'] = 'list,DialogChoice'

# 合并choices字段（假设DialogChoice有2个字段：text和next_dialog_id）
sheet.merge_cells('E1:F1')
sheet.merge_cells('E2:F2')

# 第3行 - DialogChoice子字段
sheet['A3'] = '##var'
sheet['E3'] = 'text'
sheet['F3'] = 'next_dialog_id'

# 第4行
sheet['A4'] = '##'
sheet['B4'] = 'ID'

# 数据
sheet['B5'] = 8001
sheet['C5'] = 'dialog_greeting'
sheet['D5'] = 'Hello, traveler!'
sheet['E5'] = 'Tell me about quests'
sheet['F5'] = 8002

sheet['D6'] = 'How may I help you?'
sheet['E6'] = 'Goodbye'
sheet['F6'] = 0

wb.save(file_path)
print("[OK] dialog.xlsx")

# ====================================================================================
# 2. recipe.xlsx
# ====================================================================================
print("[2/2] 重新创建recipe.xlsx...")
file_path = os.path.join(replica_test_dir, 'Datas', 'recipe.xlsx')

wb = Workbook()
sheet = wb.active

# 第1行
sheet['A1'] = '##var'
sheet['B1'] = 'id'
sheet['C1'] = 'name'
sheet['D1'] = 'output_item'
sheet['E1'] = 'output_count'
sheet['F1'] = '*materials'
sheet['I1'] = 'unlock_level'

# 第2行
sheet['A2'] = '##type'
sheet['B2'] = 'int'
sheet['C2'] = 'string'
sheet['D2'] = 'int#ref=TbItem'
sheet['E2'] = 'int#range=[1,99]'
sheet['F2'] = 'list,Material'
sheet['I2'] = 'int#range=[1,100]'

# 合并*materials字段（Material有2个字段：item_id和count）
sheet.merge_cells('F1:H1')
sheet.merge_cells('F2:H2')

# 第3行 - Material子字段
sheet['A3'] = '##var'
sheet['G3'] = 'item_id'
sheet['H3'] = 'count'

# 第4行
sheet['A4'] = '##'
sheet['B4'] = 'ID'

# 数据
sheet['B5'] = 9001
sheet['C5'] = 'recipe_iron_sword'
sheet['D5'] = 4001
sheet['E5'] = 1
sheet['G5'] = 2001
sheet['H5'] = 10
sheet['I5'] = 5

sheet['G6'] = 2002
sheet['H6'] = 5

wb.save(file_path)
print("[OK] recipe.xlsx")

print("\n[完成] dialog和recipe已重新创建")
