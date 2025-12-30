import openpyxl
from openpyxl import Workbook
import os

# 重新创建achievement.xlsx - 修正字段结构

replica_test_dir = r'e:\Learn\luban_examples\skill_tests\replica_1_test'
file_path = os.path.join(replica_test_dir, 'Datas', 'achievement.xlsx')

print("重新创建achievement.xlsx...")

wb = Workbook()
sheet = wb.active

# 第1行: ##var
# Achievement字段: id, name, desc, condition, *rewards, points
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'desc'
sheet.cell(1, 5).value = 'condition'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=8)
sheet.cell(1, 9).value = '*rewards'
sheet.cell(1, 10).value = 'points'

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'text'
sheet.cell(2, 5).value = 'AchievementCondition'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)
sheet.cell(2, 9).value = 'list,AchievementReward'
sheet.cell(2, 10).value = 'int'

# 第3行: ##var - condition子类字段 (没有*rewards的子字段定义，因为它是多行列表)
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 5).value = '$type'
sheet.cell(3, 6).value = 'monster_type'  # KillCount
sheet.cell(3, 7).value = 'count'          # KillCount, CollectItem
sheet.cell(3, 8).value = 'item_id'        # CollectItem

# 第4行: ##
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '名称'

# 数据行1: KillCount成就
row = 5
sheet.cell(row, 2).value = 10001
sheet.cell(row, 3).value = 'ach_kill_goblin'
sheet.cell(row, 4).value = 'ach_kill_goblin_desc'
sheet.cell(row, 5).value = 'KillCount'
sheet.cell(row, 6).value = 'goblin'
sheet.cell(row, 7).value = 100
sheet.cell(row, 9).value = '1001;10'  # *rewards第1行: item_id;count
sheet.cell(row, 10).value = 50

# 数据行2: *rewards第2行（属于10001成就）
row = 6
sheet.cell(row, 9).value = '1002;5'  # *rewards第2行

# 数据行3: CollectItem成就
row = 7
sheet.cell(row, 2).value = 10002
sheet.cell(row, 3).value = 'ach_collect_gold'
sheet.cell(row, 4).value = 'ach_collect_gold_desc'
sheet.cell(row, 5).value = 'CollectItem'
sheet.cell(row, 7).value = 50
sheet.cell(row, 8).value = 2002
sheet.cell(row, 9).value = '3001;1'  # *rewards
sheet.cell(row, 10).value = 100

wb.save(file_path)
print(f"[OK] 已重新创建 achievement.xlsx")
print("  - condition字段合并列5-8")
print("  - *rewards使用sep分隔符格式: 'item_id;count'")
print("  - points在列10")
