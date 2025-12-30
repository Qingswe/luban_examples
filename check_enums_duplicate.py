import openpyxl
from openpyxl import load_workbook
import os

# 检查__enums__.xlsx中的枚举定义

data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
enums_file = os.path.join(data_dir, '__enums__.xlsx')

wb = load_workbook(enums_file)
sheet = wb.active

print("__enums__.xlsx中的所有枚举定义:")
print("=" * 80)

enum_names = {}
for row_idx in range(1, sheet.max_row + 1):
    full_name = sheet.cell(row_idx, 2).value
    if full_name and full_name.strip():
        if full_name not in enum_names:
            enum_names[full_name] = []
        enum_names[full_name].append(row_idx)

print("\n枚举统计:")
for enum_name, rows in sorted(enum_names.items()):
    if len(rows) > 1:
        print(f"❌ {enum_name}: 出现在第 {rows} 行 (重复 {len(rows)} 次)")
    else:
        print(f"  {enum_name}: 行 {rows[0]}")

print("\n" + "=" * 80)
