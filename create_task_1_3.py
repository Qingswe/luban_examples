import openpyxl
from openpyxl import load_workbook
import os

# 设置路径
data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
beans_file = os.path.join(data_dir, '__beans__.xlsx')
talent_file = os.path.join(data_dir, '#Talent.xlsx')

# Task 1.3: 天赋系统表 (TbTalent)

# 1. 在 __beans__.xlsx 中添加 PassiveEffect 多态层次
print("正在更新 __beans__.xlsx...")
wb_beans = load_workbook(beans_file)
sheet_beans = wb_beans.active

# 找到最后一行
last_row = sheet_beans.max_row

# 添加 PassiveEffect 多态层次
beans_data = [
    ['', 'PassiveEffect', '', '', '', '', '被动效果基类'],
    ['', '', '', 'stat', 'string', '', '属性名'],
    ['', '', '', 'value', 'int', '', '数值'],
    ['', '', '', 'is_percent', 'bool', '', '是否百分比'],
    ['', 'StatBonus', 'PassiveEffect', '', '', '', '属性加成'],
    ['', '', '', 'stat', 'string', '', '属性名'],
    ['', '', '', 'value', 'int', '', '数值'],
    ['', '', '', 'is_percent', 'bool', '', '是否百分比'],
    ['', 'SkillEnhance', 'PassiveEffect', '', '', '', '技能强化'],
    ['', '', '', 'skill_id', 'int#ref=TbSkillTree', '', '技能ID'],
    ['', '', '', 'bonus_damage', 'int', '', '额外伤害'],
    ['', 'TalentNode', '', '', '', '', '天赋节点'],
    ['', '', '', 'id', 'int', '', 'ID'],
    ['', '', '', 'name', 'text', '', '名称'],
    ['', '', '', 'tier', 'int', '', '层级'],
    ['', '', '', '*effects', 'list,PassiveEffect', '', '效果列表'],
]

for row_data in beans_data:
    last_row += 1
    for col_idx, value in enumerate(row_data, start=1):
        sheet_beans.cell(last_row, col_idx).value = value

wb_beans.save(beans_file)
print(f"已更新 __beans__.xlsx，添加了 PassiveEffect 多态层次和 TalentNode bean")

# 2. 创建 #Talent.xlsx 数据表
print("正在创建 #Talent.xlsx...")
wb_talent = openpyxl.Workbook()
sheet = wb_talent.active

# 第1行: ##var
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'id'
sheet.cell(1, 3).value = 'name'
sheet.cell(1, 4).value = 'tier'
sheet.cell(1, 5).value = '*effects'
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=9)

# 第2行: ##type
sheet.cell(2, 1).value = '##type'
sheet.cell(2, 2).value = 'int'
sheet.cell(2, 3).value = 'text'
sheet.cell(2, 4).value = 'int'
sheet.cell(2, 5).value = 'list,PassiveEffect'
sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=9)

# 第3行: ##var (子字段)
sheet.cell(3, 1).value = '##var'
sheet.cell(3, 5).value = '$type'
sheet.cell(3, 6).value = 'stat'
sheet.cell(3, 7).value = 'value'
sheet.cell(3, 8).value = 'is_percent'
sheet.cell(3, 9).value = 'skill_id'
sheet.cell(3, 10).value = 'bonus_damage'

# 第4行: ## (注释行)
sheet.cell(4, 1).value = '##'
sheet.cell(4, 2).value = 'ID'
sheet.cell(4, 3).value = '天赋名称'
sheet.cell(4, 4).value = '层级'
sheet.cell(4, 5).value = '类型'
sheet.cell(4, 6).value = '属性'
sheet.cell(4, 7).value = '数值'
sheet.cell(4, 8).value = '百分比'
sheet.cell(4, 9).value = '技能ID'
sheet.cell(4, 10).value = '伤害加成'

# 数据行 - 天赋1: 力量强化 (多个StatBonus效果)
row = 5
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 1001
sheet.cell(row, 3).value = 'talent_strength'
sheet.cell(row, 4).value = 1
sheet.cell(row, 5).value = 'StatBonus'
sheet.cell(row, 6).value = 'atk'
sheet.cell(row, 7).value = 10
sheet.cell(row, 8).value = False

# 第二个效果 (多行列表)
row = 6
sheet.cell(row, 5).value = 'StatBonus'
sheet.cell(row, 6).value = 'hp'
sheet.cell(row, 7).value = 50
sheet.cell(row, 8).value = False

# 天赋2: 暴击强化 (百分比加成)
row = 7
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 1002
sheet.cell(row, 3).value = 'talent_critical'
sheet.cell(row, 4).value = 2
sheet.cell(row, 5).value = 'StatBonus'
sheet.cell(row, 6).value = 'crit_rate'
sheet.cell(row, 7).value = 15
sheet.cell(row, 8).value = True

# 天赋3: 技能专精 (SkillEnhance)
row = 8
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 1003
sheet.cell(row, 3).value = 'talent_skill_master'
sheet.cell(row, 4).value = 3
sheet.cell(row, 5).value = 'SkillEnhance'
sheet.cell(row, 9).value = 2001
sheet.cell(row, 10).value = 100

# 第二个效果
row = 9
sheet.cell(row, 5).value = 'StatBonus'
sheet.cell(row, 6).value = 'mp'
sheet.cell(row, 7).value = 30
sheet.cell(row, 8).value = False

# 天赋4: 混合天赋 (StatBonus + SkillEnhance)
row = 10
sheet.cell(row, 1).value = ''
sheet.cell(row, 2).value = 1004
sheet.cell(row, 3).value = 'talent_hybrid'
sheet.cell(row, 4).value = 3
sheet.cell(row, 5).value = 'StatBonus'
sheet.cell(row, 6).value = 'def'
sheet.cell(row, 7).value = 20
sheet.cell(row, 8).value = True

row = 11
sheet.cell(row, 5).value = 'SkillEnhance'
sheet.cell(row, 9).value = 2002
sheet.cell(row, 10).value = 50

wb_talent.save(talent_file)
print(f"已创建 #Talent.xlsx")
print("\nTask 1.3 完成!")
print("创建的文件:")
print(f"  - 更新了 __beans__.xlsx (添加 PassiveEffect 和 TalentNode)")
print(f"  - 创建了 #Talent.xlsx (多行列表 + 多态组合)")
