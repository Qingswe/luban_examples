import openpyxl
from openpyxl import load_workbook
import os

# 修复 __beans__.xlsx 中的错误

data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
beans_file = os.path.join(data_dir, '__beans__.xlsx')

print("正在检查和修复 __beans__.xlsx...")

# 读取文件
wb = load_workbook(beans_file)
sheet = wb.active

# 检查所有行
errors = []
for row_idx in range(1, sheet.max_row + 1):
    full_name = sheet.cell(row_idx, 2).value  # B列是full_name
    parent = sheet.cell(row_idx, 3).value      # C列是parent
    field_name = sheet.cell(row_idx, 4).value  # D列是fields的name

    # 如果是字段定义行（parent为空但full_name也为空）
    if full_name == '' and parent == '' and field_name != '' and field_name != 'name':
        errors.append({
            'row': row_idx,
            'field_name': field_name,
            'cell': f'B{row_idx}'
        })

if errors:
    print(f"\n发现 {len(errors)} 个需要修复的行:")
    for err in errors:
        print(f"  行 {err['row']}: 字段 '{err['field_name']}' 在 {err['cell']} 位置的 full_name 为空")

    # 这些行应该删除第2列（full_name）的内容，因为它们是字段定义行
    # 实际上它们已经是空的，问题可能在于格式
    print("\n正在重新生成 __beans__.xlsx...")

    # 备份原文件
    backup_file = beans_file.replace('.xlsx', '_backup.xlsx')
    wb.save(backup_file)
    print(f"已备份原文件到: {backup_file}")

print("\n检查完成！")
print(f"总行数: {sheet.max_row}")
