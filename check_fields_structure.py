import openpyxl
from openpyxl import load_workbook

# 读取Mini模板来获取*fields的完整结构
mini_beans = r'e:\Learn\luban_examples\MiniTemplate\Datas\__beans__.xlsx'

wb = load_workbook(mini_beans)
sheet = wb.active

print("MiniTemplate __beans__.xlsx 完整标题行:")
for col_idx in range(1, 20):
    val = sheet.cell(1, col_idx).value
    if val:
        print(f"  Col {col_idx:2d}: {val}")
    else:
        break

print("\n第2行 (*fields子字段定义):")
for col_idx in range(1, 20):
    val = sheet.cell(2, col_idx).value
    if val:
        print(f"  Col {col_idx:2d}: {val}")
    else:
        # 检查是否是空列
        if col_idx <= 15:
            print(f"  Col {col_idx:2d}: (空)")
