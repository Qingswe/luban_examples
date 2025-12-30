import openpyxl
from openpyxl import load_workbook
import os
import shutil

# 重新生成__beans__.xlsx，使用正确的格式

data_dir = r'e:\Learn\luban_examples\DataTables\Datas'
beans_file = os.path.join(data_dir, '__beans__.xlsx')

# 创建新文件
wb = openpyxl.Workbook()
sheet = wb.active

# 正确的标题行（参考MiniTemplate）
sheet.cell(1, 1).value = '##var'
sheet.cell(1, 2).value = 'full_name'
sheet.cell(1, 3).value = 'parent'
sheet.cell(1, 4).value = 'valueType'
sheet.cell(1, 5).value = 'sep'
sheet.cell(1, 6).value = 'alias'
sheet.cell(1, 7).value = 'comment'
sheet.cell(1, 8).value = 'group'
sheet.cell(1, 9).value = 'tags'
sheet.cell(1, 10).value = '*fields'
sheet.cell(1, 11).value = '*fields'
sheet.cell(1, 12).value = '*fields'
sheet.cell(1, 13).value = '*fields'

# 合并*fields列
sheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=16)

# 第2行: ##var子行定义
sheet.cell(2, 1).value = '##var'
sheet.cell(2, 10).value = 'name'
sheet.cell(2, 11).value = 'alias'
sheet.cell(2, 12).value = 'type'
sheet.cell(2, 13).value = 'group'
sheet.cell(2, 14).value = 'comment'
sheet.cell(2, 15).value = 'tags'
sheet.cell(2, 16).value = 'variants'

# 第3行: 注释行
sheet.cell(3, 1).value = '##'
sheet.cell(3, 2).value = '全名(带命名空间前缀)'
sheet.cell(3, 5).value = '分隔符'
sheet.cell(3, 10).value = '*fields'

current_row = 4

# 所有Bean定义（使用*fields多行列表格式）
beans = [
    # Bean定义: (full_name, parent, comment, fields_list)
    # fields_list中每项为: (name, type, group, comment)

    ('PassiveEffect', '', '被动效果基类', []),

    ('StatBonus', 'PassiveEffect', '属性加成', [
        ('stat', 'string', '', '属性名'),
        ('value', 'int', '', '数值'),
        ('is_percent', 'bool', '', '是否百分比'),
    ]),

    ('SkillEnhance', 'PassiveEffect', '技能强化', [
        ('skill_id', 'int#ref=TbSkillTree', '', '技能ID'),
        ('bonus_damage', 'int', '', '额外伤害'),
    ]),

    ('TalentNode', '', '天赋节点', [
        ('id', 'int', '', 'ID'),
        ('name', 'text', '', '名称'),
        ('tier', 'int', '', '层级'),
        ('*effects', 'list,PassiveEffect', '', '效果列表'),
    ]),

    ('Modifier', '', '属性调整器基类', []),

    ('FlatModifier', 'Modifier', '固定值调整器', [
        ('stat', 'string', '', '属性名'),
        ('value', 'int', '', '数值'),
    ]),

    ('PercentModifier', 'Modifier', '百分比调整器', [
        ('stat', 'string', '', '属性名'),
        ('percent', 'float', '', '百分比'),
    ]),

    ('Affix', '', '词缀', [
        ('id', 'int', '', 'ID'),
        ('name', 'string', '', '名称'),
        ('modifier', 'Modifier', '', '调整器'),
    ]),

    ('Equipment', '', '装备', [
        ('id', 'int', '', 'ID'),
        ('name', 'string', '', '名称'),
        ('slot', 'EquipSlot', '', '装备槽位'),
        ('base_stats', 'map,string,int', '', '基础属性'),
        ('affix_pool', '(list#index=id),Affix', '', '词缀池'),
        ('set_id', 'int?', '', '套装ID'),
    ]),

    ('Material', '', '材料', [
        ('item_id', 'int#ref=TbItem', '', '物品ID'),
        ('count', 'int#range=[1,999]', '', '数量'),
    ]),

    ('Recipe', '', '配方', [
        ('id', 'int', '', 'ID'),
        ('name', 'string', '', '名称'),
        ('output_item', 'int#ref=TbItem', '', '输出物品'),
        ('output_count', 'int#range=[1,99]', '', '输出数量'),
        ('*materials', 'list,Material', '', '材料列表'),
        ('unlock_level', 'int#range=[1,100]', '', '解锁等级'),
    ]),

    ('Effect', '', '效果基类', []),
    ('DamageEffect', 'Effect', '伤害效果', []),

    ('PhysicalDamage', 'DamageEffect', '物理伤害', [
        ('base_damage', 'int', '', '基础伤害'),
        ('armor_pen', 'float', '', '护甲穿透'),
    ]),

    ('MagicalDamage', 'DamageEffect', '魔法伤害', [
        ('base_damage', 'int', '', '基础伤害'),
        ('element', 'Element', '', '元素'),
        ('ignore_resist', 'bool', '', '无视抗性'),
    ]),

    ('HealEffect', 'Effect', '治疗效果', [
        ('heal_base', 'int', '', '基础治疗'),
        ('heal_scale', 'float', '', '治疗系数'),
    ]),

    ('SkillEffectConfig', '', '技能效果配置', [
        ('id', 'int', '', 'ID'),
        ('name', 'string', '', '名称'),
        ('effect', 'Effect', '', '效果'),
        ('scaling_stat', 'string', '', '成长属性'),
        ('scaling_factor', 'float', '', '成长系数'),
    ]),

    ('BuffEffect', '', 'Buff效果基类', []),

    ('StatModifier', 'BuffEffect', '属性调整', [
        ('stat', 'string', '', '属性名'),
        ('value', 'int', '', '数值'),
        ('is_percent', 'bool', '', '是否百分比'),
    ]),

    ('DotEffect', 'BuffEffect', '持续伤害', [
        ('damage_per_tick', 'int', '', '每跳伤害'),
        ('tick_interval', 'float', '', '跳动间隔'),
    ]),

    ('ControlEffect', 'BuffEffect', '控制效果', [
        ('control_type', 'string', '', '控制类型'),
    ]),

    ('Buff', '', 'Buff', [
        ('id', 'int', '', 'ID'),
        ('name', 'text', '', '名称'),
        ('buff_type', 'BuffType', '', 'Buff类型'),
        ('duration', 'float#range=[0,3600]', '', '持续时间'),
        ('stack_limit', 'int#range=[1,99]', '', '叠加上限'),
        ('refresh_on_apply', 'bool', '', '重新施加时刷新'),
        ('effect', 'BuffEffect', '', '效果'),
    ]),

    ('ElementRelation', '', '元素克制关系', [
        ('source_element', 'Element', '', '源元素'),
        ('damage_multipliers', 'map,Element,float', '', '伤害倍率'),
    ]),

    ('Quest', '', '任务', [
        ('id', 'int', '', 'ID'),
        ('title', 'text', '', '标题'),
        ('desc', 'text', '', '描述'),
        ('quest_type', 'QuestType', '', '任务类型'),
        ('prereq_quest', 'int#ref=TbQuest?', '', '前置任务'),
        ('reward_gold', 'int', '', '金币奖励'),
        ('reward_exp', 'int', '', '经验奖励'),
    ]),

    ('DialogChoice', '', '对话选项基类', []),

    ('SimpleChoice', 'DialogChoice', '简单选项', [
        ('text', 'text', '', '文本'),
        ('next_dialog', 'int#ref=TbDialog?', '', '下一对话'),
    ]),

    ('ConditionChoice', 'DialogChoice', '条件选项', [
        ('text', 'text', '', '文本'),
        ('condition', 'string', '', '条件'),
        ('next_dialog', 'int#ref=TbDialog?', '', '下一对话'),
    ]),

    ('ActionChoice', 'DialogChoice', '动作选项', [
        ('text', 'text', '', '文本'),
        ('action', 'string', '', '动作'),
        ('params', 'string', '', '参数'),
    ]),

    ('DialogLine', '', '对话行', [
        ('speaker', 'text', '', '说话者'),
        ('content', 'text', '', '内容'),
    ]),

    ('Dialog', '', '对话', [
        ('id', 'int', '', 'ID'),
        ('name', 'string', '', '名称'),
        ('*lines', 'list,DialogLine', '', '对话行'),
        ('choices', 'list,DialogChoice', '', '选项'),
    ]),

    ('AchievementCondition', '', '成就条件基类', []),

    ('KillCount', 'AchievementCondition', '击杀数量', [
        ('monster_type', 'string', '', '怪物类型'),
        ('count', 'int', '', '数量'),
    ]),

    ('CollectItem', 'AchievementCondition', '收集物品', [
        ('item_id', 'int#ref=TbItem', '', '物品ID'),
        ('count', 'int', '', '数量'),
    ]),

    ('CompleteQuest', 'AchievementCondition', '完成任务', [
        ('quest_id', 'int#ref=TbQuest', '', '任务ID'),
    ]),

    ('ReachLevel', 'AchievementCondition', '达到等级', [
        ('level', 'int#range=[1,100]', '', '等级'),
    ]),

    ('AchievementReward', '', '成就奖励', [
        ('item_id', 'int#ref=TbItem', '', '物品ID'),
        ('count', 'int#range=[1,999]', '', '数量'),
    ]),

    ('Achievement', '', '成就', [
        ('id', 'int', '', 'ID'),
        ('name', 'text', '', '名称'),
        ('desc', 'text', '', '描述'),
        ('condition', 'AchievementCondition', '', '条件'),
        ('*rewards', 'list,AchievementReward', '', '奖励'),
        ('points', 'int', '', '成就点数'),
    ]),
]

# 写入所有Bean定义
for full_name, parent, comment, fields in beans:
    # Bean定义行
    sheet.cell(current_row, 1).value = ''  # ##var列为空
    sheet.cell(current_row, 2).value = full_name
    sheet.cell(current_row, 3).value = parent
    sheet.cell(current_row, 7).value = comment
    current_row += 1

    # 字段定义行（*fields多行列表）
    for field_name, field_type, field_group, field_comment in fields:
        sheet.cell(current_row, 1).value = ''  # ##var列为空
        sheet.cell(current_row, 2).value = ''  # full_name为空
        sheet.cell(current_row, 3).value = ''  # parent为空
        sheet.cell(current_row, 10).value = field_name   # *fields的name
        sheet.cell(current_row, 11).value = ''  # alias
        sheet.cell(current_row, 12).value = field_type   # type
        sheet.cell(current_row, 13).value = field_group  # group
        sheet.cell(current_row, 14).value = field_comment  # comment
        sheet.cell(current_row, 15).value = ''  # tags
        sheet.cell(current_row, 16).value = ''  # variants
        current_row += 1

# 保存文件
wb.save(beans_file)
print(f"已重新生成 __beans__.xlsx (正确格式)")
print(f"总共定义了 {len(beans)} 个Bean")
print(f"总行数: {current_row - 1}")
